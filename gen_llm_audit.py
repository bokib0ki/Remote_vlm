#!/usr/bin/env python3
"""
LLM 审核表格生成器（跟 VQA_NLI_Audit 一样的结构 + 嵌入图片）

读所有 lever_test2/**/vqa_l3_llm.json，按 llm_correct_l3 排序（1 在前 0 在后），
生成审核表，每行嵌入对应 image_id 的图片（80x80 缩略图）。

列结构（15 列 + 1 列 image）：
  1. judge          ✓ 判对 / ✗ 判错
  2. correct        0/1
  3. model
  4. mode
  5. k
  6. _idx
  7. image_id
  8. type
  9. LLM_response   '1' 或 '0'（LLM 输出）
  10. short_circuit  'llm_judge'（LLM 4 步短路）
  11. ok             bool
  12. question
  13. gt
  14. pred
  15. (空，备用)
  + image 列（嵌入 80x80 缩略图）

用法:
  python gen_llm_audit.py
  python gen_llm_audit.py --max-rows 2000  # 限制行数
  python gen_llm_audit.py --thumb-size 60   # 缩略图尺寸
"""
import argparse
import io
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

RAW_ROOT = Path('/home/admin1/models/raw_outputs/lever_test2')
VQA3     = Path('/home/admin1/models/raw_outputs/vqa_3sets')
OUT_PATH = Path('/home/admin1/projects/remote_vlm_eval/excel_result/VQA_LLM_Audit.xlsx')
IMG_DIR  = Path('/home/admin1/models/vrsbench_images/Images_val')
VRS_PATH = Path('/home/admin1/models/VRSBench_EVAL_vqa.json')


def parse_dir_name(dirname: str):
    """识别 lever_test2 / vqa_3sets 两种 pattern"""
    for pat in [
        r'^(.+?)_(think(?:ON|OFF))_lever_test2_lever_k=(\d+)_vqa$',
        r'^(.+?)_(think(?:ON|OFF))_vqa_3sets_lever_k=(\d+)_vqa_(\d+)$',
    ]:
        m = re.match(pat, dirname)
        if m:
            if 'vqa_3sets' in pat:
                return m.group(1), m.group(2), int(m.group(3)), int(m.group(4))
            return m.group(1), m.group(2), int(m.group(3)), None
    return None


def load_type_map():
    """_idx -> type"""
    ann = json.load(open(VRS_PATH))
    tmap = {}
    for it in ann:
        if 'question_id' in it:
            tmap[it['question_id']] = it.get('type', '<UNKNOWN>')
        elif '_idx' in it:
            tmap[it['_idx']] = it.get('type', '<UNKNOWN>')
    return tmap


def make_thumb(image_id, thumb_size, cache_dir):
    """
    把图片缩到 thumb_size, 保存到 cache_dir, 返回路径。
    缩略图缓存到 cache_dir 复用。
    """
    if not image_id:
        return None
    src = IMG_DIR / image_id
    if not src.exists():
        return None
    cache_path = cache_dir / f'{image_id}'
    if cache_path.exists():
        return cache_path
    try:
        img = PILImage.open(src).convert('RGB')
        img.thumbnail((thumb_size, thumb_size), PILImage.LANCZOS)
        img.save(cache_path, 'PNG', optimize=True)
        return cache_path
    except Exception:
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--max-rows', type=int, default=0,
                    help='限制行数（0=全量）')
    ap.add_argument('--thumb-size', type=int, default=80,
                    help='缩略图边长（默认 80 像素）')
    ap.add_argument('--include-vqa3sets', action='store_true',
                    help='是否包含 vqa_3sets/ 下的 18 个新 run')
    ap.add_argument('--batch-dir', type=Path, default=None,
                    help='指定扫的根目录（覆盖默认 lever_test2）')
    args = ap.parse_args()

    raw_root = args.batch_dir or RAW_ROOT
    cache_dir = Path(f'/tmp/vqa_llm_audit_thumbs_{args.thumb_size}')
    cache_dir.mkdir(parents=True, exist_ok=True)

    # 1. 收集所有 vqa_l3_llm.json
    search_paths = [raw_root]
    if args.include_vqa3sets and args.batch_dir is None:
        search_paths.append(VQA3)

    all_records = []
    n_files = 0
    n_imgs_ok = 0
    n_imgs_missing = 0

    type_map = load_type_map()

    for sp in search_paths:
        for l3f in sorted(sp.glob('**/vqa_l3_llm.json')):
            parsed = parse_dir_name(l3f.parent.name)
            if not parsed:
                print(f'[WARN] 跳过（目录名不匹配）: {l3f.parent.name}')
                continue
            model, mode, k, vqa3_n = parsed
            n_files += 1
            try:
                recs = json.load(open(l3f))
            except Exception as e:
                print(f'[ERR] {l3f}: {e}')
                continue
            for r in recs:
                image_id = r.get('image_id', '')
                # 缩略图预生成（后续 add image 时直接引用）
                if image_id and (IMG_DIR / image_id).exists():
                    make_thumb(image_id, args.thumb_size, cache_dir)
                    n_imgs_ok += 1
                else:
                    n_imgs_missing += 1
                all_records.append({
                    'model':     model,
                    'mode':      mode,
                    'k':         k,
                    'vqa3_n':    vqa3_n,
                    '_idx':      r.get('_idx', ''),
                    'image_id':  image_id,
                    'type':      type_map.get(r.get('_idx', ''), '<UNKNOWN>'),
                    'question':  r.get('question', ''),
                    'gt':        r.get('gt', ''),
                    'pred':      r.get('pred', ''),
                    'llm_response':   r.get('llm_response', ''),
                    'short_circuit':  r.get('llm_short_circuit', 'llm_judge'),
                    'ok':        r.get('llm_ok', True),
                    'llm_correct_l3': int(r.get('llm_correct_l3', 0)),
                })

    # 排序：correct 1 在前，相同按 llm_response 排
    all_records.sort(key=lambda r: (-r['llm_correct_l3'], r.get('llm_response', '')))

    # 限制行数
    if args.max_rows > 0 and len(all_records) > args.max_rows:
        # 取前 max_rows (判对 + 判错各半)
        n_correct = sum(1 for r in all_records if r['llm_correct_l3'] == 1)
        n_wrong = len(all_records) - n_correct
        take_c = min(n_correct, args.max_rows // 2 + 100)
        take_w = args.max_rows - take_c
        records_correct = [r for r in all_records if r['llm_correct_l3'] == 1][:take_c]
        records_wrong = [r for r in all_records if r['llm_correct_l3'] == 0][:take_w]
        all_records = records_correct + records_wrong
        print(f'[限制] 取前 {args.max_rows} 行（判对 {len(records_correct)} + 判错 {len(records_wrong)}）')

    n_total = len(all_records)
    n_correct = sum(1 for r in all_records if r['llm_correct_l3'] == 1)
    n_wrong = n_total - n_correct
    print(f'\n扫描: {n_files} 个 vqa_l3_llm.json')
    print(f'记录: {n_total} 条 (correct={n_correct}, wrong={n_wrong}, acc={n_correct/n_total*100:.2f}%)')
    print(f'图片: {n_imgs_ok} OK, {n_imgs_missing} 缺失')

    # 2. 写 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'LLM_Audit'

    HEADER_FILL = PatternFill("solid", fgColor='305496')
    BOLD_WHITE  = Font(bold=True, color='FFFFFF', size=11)
    BORDER      = Border(*[Side(style='thin', color='BFBFBF')] * 4)
    CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT        = Alignment(horizontal='left', vertical='center', wrap_text=True)

    CORRECT_FILL = PatternFill("solid", fgColor='C6EFCE')
    WRONG_FILL   = PatternFill("solid", fgColor='FFC7CE')

    headers = [
        'judge', 'correct', 'model', 'mode', 'k', '_idx', 'image_id', 'type',
        'LLM_response', 'short_circuit', 'ok',
        'question', 'gt', 'pred',
        'image',  # 最后一列
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HEADER_FILL
        c.font = BOLD_WHITE
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 28

    for ri, r in enumerate(all_records, 2):
        is_c = r['llm_correct_l3'] == 1
        judge_disp = '✓ 判对' if is_c else '✗ 判错'
        fill = CORRECT_FILL if is_c else WRONG_FILL

        cells = [
            (judge_disp,                 CENTER),
            (r['llm_correct_l3'],        CENTER),
            (r['model'],                 LEFT),
            (r['mode'],                  CENTER),
            (r['k'],                     CENTER),
            (r['_idx'],                  CENTER),
            (r['image_id'],              LEFT),
            (r['type'],                  CENTER),
            (r['llm_response'][:50],     LEFT),
            (r['short_circuit'],         CENTER),
            ('✓' if r['ok'] else '✗',    CENTER),
            (r['question'],              LEFT),
            (r['gt'],                    LEFT),
            (r['pred'],                  LEFT),
        ]
        for ci, (v, align) in enumerate(cells, 1):
            c = ws.cell(ri, ci, v)
            c.fill = fill if ci <= 2 else PatternFill("solid", fgColor='FFFFFF' if ri % 2 == 0 else 'F2F2F2')
            c.font = Font(bold=(ci <= 2), size=10, color='1F1F1F')
            c.alignment = align
            c.border = BORDER

        # 嵌入缩略图（第 15 列）
        image_id = r['image_id']
        if image_id:
            thumb_path = cache_dir / image_id
            if thumb_path.exists():
                try:
                    img = XLImage(str(thumb_path))
                    # 缩略图本身已经是 80x80（或 thumb_size），不再缩
                    anchor = f'O{ri}'  # O 列是第 15 列
                    img.anchor = anchor
                    ws.add_image(img)
                except Exception as e:
                    pass

        # 行高跟缩略图大小一致
        ws.row_dimensions[ri].height = max(args.thumb_size + 4, 24)

    # 列宽
    widths = [10, 8, 22, 10, 6, 8, 18, 22, 14, 14, 6, 50, 18, 30, args.thumb_size * 0.9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'C2'  # 冻结 judge + correct + 后续滚动

    # 顶部加统计行
    ws.insert_rows(1)
    stat = f'LLM L3 评测审核表（Qwen3.5-4B judge, 温度=0）| 共 {n_total} 条 | ✓ 判对 {n_correct} ({n_correct/n_total*100:.2f}%) | ✗ 判错 {n_wrong} ({n_wrong/n_total*100:.2f}%) | 图片缩略图 {args.thumb_size}px'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(1, 1, stat)
    c.fill = HEADER_FILL
    c.font = Font(bold=True, color='FFFFFF', size=12)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))
    print(f'\n[完成] 写入 {OUT_PATH}')
    print(f'      Sheet: LLM_Audit ({n_total+2} 行 × {len(headers)} 列)')

    # 按 type + correct 统计
    print('\n=== 按 type 统计 ===')
    type_stat = {}
    for r in all_records:
        t = r['type']
        if t not in type_stat:
            type_stat[t] = {'n': 0, 'c': 0}
        type_stat[t]['n'] += 1
        type_stat[t]['c'] += r['llm_correct_l3']
    for t, s in sorted(type_stat.items(), key=lambda x: -x[1]['n']):
        acc = s['c']/s['n']*100 if s['n'] else 0
        print(f'  {t:25}  N={s["n"]:>5}  correct={s["c"]:>4}  acc={acc:>5.2f}%')


if __name__ == '__main__':
    main()
