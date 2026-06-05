#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_vqa_type_report_v3.py
==========================
VRS-VQA 分层评测报告生成器（v3.0）
处理 eval_select.py 输出的 VRS-VQA raw 数据，生成多模型对比 Excel 报告。

──────────────────────────────────────────────────────────────
重要：VRSBench VQA 三级评测机制
──────────────────────────────────────────────────────────────
eval_select.py 跑完后，每条记录有 judge_level + correct 字段:
  L1: substring 匹配（gt in pred, 大小写无关）        → correct=1/0
  L2: yes/no/数字 精确匹配                             → correct=1/0
  L3: Qwen3.7 / GPT-4o-mini 语义匹配（待评测）         → correct=0 (pending)

L3 的题需要另外跑 VQA_V3.py（调用 LLM API）才能得到真 correct。

本脚本的 acc 计算口径:
  acc = (L1_correct + L2_correct) / 总 N
  → L3 pending 题的 correct 强制按 0 算（即"无 L3 分析结果就当错"）
  → 这是因为如果不跑 L3，模型在 L3 题上没有任何"已知正确"的证据，
     计入分母才能反映真实准确率，而不是虚高。

──────────────────────────────────────────────────────────────
用法
──────────────────────────────────────────────────────────────
默认:
  python3 gen_vqa_type_report_v3.py

自定义:
  python3 gen_vqa_type_report_v3.py --raw-root /path/to/raw --out-dir /path/to/xlsx
  python3 gen_vqa_type_report_v3.py --name my_report --l3-as-wrong false
  python3 gen_vqa_type_report_v3.py --include-mode thinkOFF

参数:
  --raw-root     扫 raw_outputs 根目录（默认: /home/admin1/models/raw_outputs/lever_test2）
  --out-dir      xlsx 输出目录（默认: /home/admin1/projects/remote_vlm_eval/excel_result）
  --name         xlsx 文件名（不含 .xlsx 后缀，默认: VQA_type_report_v3）
  --include-mode 只统计指定模式（默认: 全部）
  --l3-as-wrong  L3 pending 是否按 0 算（默认: false → 2026-06-03 改用整体 (L1-L3) acc）
                 true:  acc = (L1_c + L2_c + 0*L3) / N_total   ← 旧口径（L3 算错）
                 false: acc = (L1_c + L2_c + L3_c) / N_total   ← 当前默认（整体 (L1-L3)）

输出 3 个 Sheet:
  1. per_lever  - 每 (model, mode, k) 行的 N / perf / (L1-L3) 整体 acc / (L1-L3) 排名
                 / L3 N / L3 正确数 / L3 子集 acc / L3 子集 排名
  2. type_acc   - 12 type × (model, k) 转置矩阵：行=k，列=type，含 (L1-L3) 整体 acc + 排名
  3. raw_concat - 全部原始记录扁平

期望的 raw_outputs 目录结构:
  /raw_root/lever_k=20_vqa/
    {model}_{mode}_lever_test2_lever_k=20_vqa/raw_outputs.json
    ...
"""
import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 默认路径（可被 CLI 覆盖） ──────────────────────────────
DEFAULT_RAW_ROOT = Path('/home/admin1/models/raw_outputs/lever_test2')
DEFAULT_OUT_DIR  = Path('/home/admin1/projects/remote_vlm_eval/excel_result')
VQA_ANNOT        = Path('/home/admin1/models/VRSBench_EVAL_vqa.json')

# ─── 12 个 VRS-VQA 任务类型（固定顺序） ─────────────────────
TYPE_ORDER = [
    'object quantity', 'object existence', 'object position',
    'object category', 'object color', 'object shape',
    'object size', 'object direction', 'scene type',
    'rural or urban', 'reasoning', 'image',
]
TYPE_CN = {
    'object quantity':  '物体计数',
    'object existence': '物体存在',
    'object position':  '物体位置',
    'object category':  '物体类别',
    'object color':     '物体颜色',
    'object shape':     '物体形状',
    'object size':      '物体大小',
    'object direction': '物体方向',
    'scene type':       '场景类型',
    'rural or urban':   '城乡分类',
    'reasoning':        '推理',
    'image':            '图像属性',
}

# ─── 模型显示名 + 底色（按 model 唯一） ────────────────────
MODEL_DISPLAY = {
    'qwen2.5-vl-3B':         'Qwen2.5-VL-3B',
    'qwen2.5-vl-7B':         'Qwen2.5-VL-7B',
    'qwen3.5-0.8B':          'Qwen3.5-0.8B',
    'qwen3.5-2B':            'Qwen3.5-2B',
    'qwen3.5-4B':            'Qwen3.5-4B',
    'qwen3-vl-2B':           'Qwen3-VL-2B',
    'qwen3-vl-4B':           'Qwen3-VL-4B',
    'qwen3-vl-4B-thinking':  'Qwen3-VL-4B-Thinking',
    'minicpm-v-4.6':         'MiniCPM-V 4.6',
    'gemma-4-e2b':           'Gemma-4-2B',
    'gemma-4-e4b':           'Gemma-4-4B',
}
MODEL_COLOR = {
    'qwen2.5-vl-3B':         'D6EAF8',
    'qwen2.5-vl-7B':         '85C1E2',
    'qwen3.5-0.8B':          'FCF3CF',
    'qwen3.5-2B':            'F9E79F',
    'qwen3.5-4B':            'F7DC6F',
    'qwen3-vl-2B':           'A2D9CE',
    'qwen3-vl-4B':           '76D7C4',
    'qwen3-vl-4B-thinking':  'FAD7A0',
    'minicpm-v-4.6':         'D5F5E3',
    'gemma-4-e2b':           'F5CBA7',
    'gemma-4-e4b':           'EDBB99',
}
# 排名配色（rank 1=金, 2=银, 3=铜）
RANK_COLOR = {1: 'FFD700', 2: 'C0C0C0', 3: 'CD7F32'}

# ─── 通用样式 ───────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL   = PatternFill("solid", fgColor="17202A")
TOTAL_FILL   = PatternFill("solid", fgColor="EAEDED")
BOLD_WHITE   = Font(bold=True, color="FFFFFF", size=10)
BOLD_DARK    = Font(bold=True, color="1F1F1F", size=10)
BOLD_TITLE   = Font(bold=True, color="FFFFFF", size=13)
NORMAL       = Font(size=10)
THIN         = Side(style='thin')
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER       = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT         = Alignment(horizontal='left',   vertical='center', wrap_text=False)


# ─── 工具函数 ───────────────────────────────────────────────
def parse_dir(dirname: str):
    """解析 '{model}_{mode}_lever_test2_lever_k={k}_vqa' 目录名 → (model, mode, k)"""
    for pat in [
        r'^(.+?)_(think(?:ON|OFF))_lever_test2_lever_k=(\d+)_vqa$',
        r'^(.+?)_(think(?:ON|OFF))_lever_k=(\d+)$',
    ]:
        m = re.match(pat, dirname)
        if m:
            return m.group(1), m.group(2), int(m.group(3))
    return None


def discover_runs(raw_root: Path, include_modes: set | None = None):
    """扫 raw_root/lever_k=*_vqa/ 下所有 run。返回 [(model, mode, k, raw_path), ...]"""
    runs = []
    if not raw_root.exists():
        return runs
    for lever_dir in sorted(raw_root.iterdir()):
        if not lever_dir.is_dir():
            continue
        for run_dir in sorted(lever_dir.iterdir()):
            if not run_dir.is_dir():
                continue
            parsed = parse_dir(run_dir.name)
            if parsed is None:
                continue
            model, mode, k = parsed
            if include_modes and mode not in include_modes:
                continue
            raw_file = run_dir / 'raw_outputs.json'
            if raw_file.exists():
                runs.append((model, mode, k, raw_file))
    return runs


def load_type_map():
    with open(VQA_ANNOT) as f:
        vrs = json.load(f)
    return {str(it['question_id']): it['type'] for it in vrs}


def fmt_pct(c, n):
    return f'{c/n:.2%}' if n > 0 else '—'


def acc_color(acc):
    if acc is None:
        return '808080'
    if acc >= 0.90: return '0B5345'
    if acc >= 0.70: return '1A5276'
    if acc >= 0.50: return 'B9770E'
    return '922B21'


def rank_models(acc_dict: dict, descending: bool = True) -> dict:
    """给 {model_key: acc|None} 算排名。None 不参与。降序时 acc 高→rank 1。平局取最小 rank。"""
    valid = {k: v for k, v in acc_dict.items() if v is not None}
    if not valid:
        return {k: None for k in acc_dict}
    sorted_items = sorted(valid.items(), key=lambda x: x[1], reverse=descending)
    rank = {}
    last_v, last_rank = None, 0
    for i, (k, v) in enumerate(sorted_items, 1):
        if v != last_v:
            last_rank = i
            last_v = v
        rank[k] = last_rank
    for k in acc_dict:
        if k not in rank:
            rank[k] = None
    return rank


def fill_cell(cell, value=None, fill=None, font=None, align=None, border=True):
    if value is not None:
        cell.value = value
    if fill:  cell.fill = fill
    if font:  cell.font = font
    if align: cell.alignment = align
    if border: cell.border = BORDER


# ─── 数据收集 ──────────────────────────────────────────────
def collect_summary(runs, type_map, l3_as_wrong: bool = True):
    """
    返回 (summary, all_raw_rows):
      summary[(model, mode, k)] = {
        'overall': {
          'total': N,           # 总题数
          'l1': n, 'l1_c': c,   # L1
          'l2': n, 'l2_c': c,   # L2
          'l3_pending': n,      # L3 pending（待评测）
          'l3_total': n,        # L3 总数（=l3_pending，如果 vqa_l3_bert.json 存在）
          'l3_correct': c,      # L3 正确数
          'l3_acc': c/n,        # L3 准确率 = l3_correct / l3_total
          'correct_total': c,   # L1_c + L2_c + (0 if l3_as_wrong else L3_c=未评)
          'acc': c/N            # 准确率（L3 算错）
        },
        'by_type': {type: 同结构, but 按 type 维度},
        'perf': {'out_avg':, 'in_avg':, 'img_avg':, 'speed_avg':}
      }
    """
    summary = {}
    all_raw_rows = []

    for model, mode, k, raw_path in runs:
        recs = json.load(open(raw_path))

        # ─── L3 评测数据（优先 LLM, 然后 NLI, 最后 BERTScore） ────
        # 路径: raw_outputs/<run>/vqa_l3_{llm,nli,bert}.json
        l3_llm_path  = raw_path.parent / 'vqa_l3_llm.json'
        l3_nli_path  = raw_path.parent / 'vqa_l3_nli.json'
        l3_bert_path = raw_path.parent / 'vqa_l3_bert.json'
        l3_idx_to_correct = {}  # _idx -> correct_l3
        l3_loaded = False
        l3_method = 'none'  # 'llm' / 'nli' / 'bertscore' / 'none'
        for l3_path, method in [(l3_llm_path, 'llm'), (l3_nli_path, 'nli'), (l3_bert_path, 'bertscore')]:
            if l3_path.exists():
                try:
                    l3_recs = json.load(open(l3_path))
                    # 兼容字段名: llm_correct_l3 / correct_l3
                    for lr in l3_recs:
                        idx = str(lr.get('_idx', ''))
                        c = lr.get('correct_l3', lr.get('llm_correct_l3', 0))
                        l3_idx_to_correct[idx] = int(c)
                    l3_loaded = True
                    l3_method = method
                    break
                except Exception as e:
                    print(f'[WARN] 读 {l3_path} 失败: {e}')

        by_type = defaultdict(lambda: {
            'total': 0, 'l1': 0, 'l1_c': 0, 'l2': 0, 'l2_c': 0,
            'l3_pending': 0, 'l3_correct': 0,
            'correct_total': 0,
        })
        overall = {
            'total': 0, 'l1': 0, 'l1_c': 0, 'l2': 0, 'l2_c': 0,
            'l3_pending': 0, 'l3_correct': 0,
            'correct_total': 0,
        }
        perf = {'out': [], 'in': [], 'img': [], 'speed': []}

        for r in recs:
            idx = str(r.get('_idx', ''))
            t = type_map.get(idx, '<UNKNOWN>')
            c = r.get('correct', 0)
            is_c = (str(c) in ('1', 'True', 'true')) or (c is True)
            lvl = r.get('judge_level', '')

            overall['total'] += 1
            bt = by_type[t]
            bt['total'] += 1

            if lvl == 'L1':
                overall['l1'] += 1
                overall['l1_c'] += int(is_c)
                bt['l1'] += 1
                bt['l1_c'] += int(is_c)
            elif lvl == 'L2':
                overall['l2'] += 1
                overall['l2_c'] += int(is_c)
                bt['l2'] += 1
                bt['l2_c'] += int(is_c)
            elif lvl == 'L3':
                # L3 pending
                overall['l3_pending'] += 1
                bt['l3_pending'] += 1
                # 如果 L3 评测结果有
                if l3_loaded and idx in l3_idx_to_correct:
                    l3_c = l3_idx_to_correct[idx]
                    overall['l3_correct'] += l3_c
                    bt['l3_correct'] += l3_c
            # 注: L1 c + L2 c 直接累加到 correct_total；L3 不加（按错算）

            overall['correct_total'] = overall['l1_c'] + overall['l2_c']
            bt['correct_total'] = bt['l1_c'] + bt['l2_c']

            # perf 字段
            perf['out'].append(r.get('tokens', 0))
            perf['in'].append(r.get('input_tokens', 0))
            perf['img'].append(r.get('img_tokens', 0))
            perf['speed'].append(r.get('speed', 0))

            all_raw_rows.append({
                'model':     model,
                'mode':      mode,
                'k':         k,
                '_idx':      idx,
                'type':      t,
                'correct':   int(is_c),
                'judge_level': lvl,
                'question':  r.get('question', ''),
                'gt':        r.get('gt', ''),
                'pred':      r.get('pred', ''),
                'image_id':  r.get('image_id', ''),
                'tokens':    r.get('tokens', 0),
                'input_tokens': r.get('input_tokens', 0),
                'img_tokens': r.get('img_tokens', 0),
                'speed':     r.get('speed', 0),
            })

        # L3 准确率 = l3_correct / l3_pending（如果 L3 评测结果有）
        if l3_loaded:
            overall['l3_total'] = overall['l3_pending']
            overall['l3_acc'] = (overall['l3_correct'] / overall['l3_total']) if overall['l3_total'] else 0
            for t_name, bt_v in by_type.items():
                bt_v['l3_total'] = bt_v['l3_pending']
                bt_v['l3_acc'] = (bt_v['l3_correct'] / bt_v['l3_total']) if bt_v['l3_total'] else 0
        else:
            overall['l3_total'] = 0
            overall['l3_acc'] = None
            for t_name, bt_v in by_type.items():
                bt_v['l3_total'] = 0
                bt_v['l3_acc'] = None

        # ─── 整体 acc（L1+L2+L3, 2026-06-03 改为整体）───
        # 整体 acc = (L1_c + L2_c + L3_c) / N_total
        # l3_as_wrong 为 True 时仍用旧口径 (L3 pending 算 0)，False 时用整体
        if l3_as_wrong:
            overall['acc'] = overall['correct_total'] / overall['total'] if overall['total'] else 0
            for t_name, bt_v in by_type.items():
                bt_v['acc'] = bt_v['correct_total'] / bt_v['total'] if bt_v['total'] else 0
        else:
            # 整体 (L1-L3) 准确率
            overall['acc'] = (overall['correct_total'] + overall['l3_correct']) / overall['total'] if overall['total'] else 0
            for t_name, bt_v in by_type.items():
                bt_v['acc'] = (bt_v['correct_total'] + bt_v.get('l3_correct', 0)) / bt_v['total'] if bt_v['total'] else 0

        # perf 平均
        n = len(perf['out'])
        perf_summary = {
            'out_avg':  sum(perf['out'])  / n if n else 0,
            'in_avg':   sum(perf['in'])   / n if n else 0,
            'img_avg':  sum(perf['img'])  / n if n else 0,
            'speed_avg': sum(perf['speed']) / n if n else 0,
            'n': n,
        }

        summary[(model, mode, k)] = {
            'overall': overall,
            'by_type': dict(by_type),
            'perf': perf_summary,
            'has_l3': l3_loaded,
            'l3_method': l3_method,
        }
    return summary, all_raw_rows


# ─── 写 Excel ──────────────────────────────────────────────
def write_excel(summary, all_raw_rows, out_path: Path, l3_as_wrong: bool):
    model_modes = sorted(set((k[0], k[1]) for k in summary.keys()))
    k_values    = sorted(set(k[2] for k in summary.keys()))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ====================================================
    # Sheet 1: per_lever（含 perf + rank，不显示 L1/L2/L3 分级）
    # ====================================================
    ws1 = wb.create_sheet('per_lever')
    headers1 = ['模型', '模式', 'k', '总题数 N', '平均输出 token', '平均速度 (t/s)',
                '平均输入 token', '平均图片 token', '准确率 (L1-L3)', '(L1-L3) 排名',
                'L3 N', 'L1+L2 准确率', 'L1+L2 排名']
    # 统计 L3 评测方法（用于标题显示）
    methods_used = sorted(set(s.get('l3_method', 'none') for s in summary.values()))
    method_label = '/'.join({'llm':'LLM', 'nli':'NLI', 'bertscore':'BERTScore', 'none':'无'}.get(m, m) for m in methods_used)
    method_note = f' | L3 评测: {method_label}' if methods_used and methods_used != ['none'] else ''

    ncol1 = len(headers1)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol1)
    l3_note = 'L3 pending 按 0 算入分母' if l3_as_wrong else '(L1+L2+L3) / N_total 整体准确率'
    fill_cell(ws1['A1'], f'VRS-VQA per_lever 总览（{l3_note}{method_note}）',
              TITLE_FILL, BOLD_TITLE, CENTER)
    ws1.row_dimensions[1].height = 28
    for ci, h in enumerate(headers1, 1):
        fill_cell(ws1.cell(2, ci, h), h, HEADER_FILL, BOLD_WHITE, CENTER)
    ws1.row_dimensions[2].height = 30

    row = 3
    for model, mode in model_modes:
        mfill  = PatternFill("solid", fgColor=MODEL_COLOR.get(model, 'FFFFFF'))
        section_start = row

        for k in k_values:
            d = summary.get((model, mode, k))
            ws1.cell(row, 1, MODEL_DISPLAY.get(model, model))
            ws1.cell(row, 2, mode)
            fill_cell(ws1.cell(row, 3, k), None, mfill, BOLD_DARK, CENTER)

            if d is None:
                for ci in range(4, ncol1 + 1):
                    fill_cell(ws1.cell(row, ci, '—'), None, mfill, NORMAL, CENTER)
            else:
                o = d['overall']
                p = d['perf']
                has_l3 = d.get('has_l3') and o.get('l3_total', 0) > 0
                l3_total_nli = o.get('l3_total', 0) if has_l3 else 0
                l3_pending = o.get('l3_pending', 0)

                # 算本 k 维度下所有模型 acc → rank（(L1-L3) 整体排名）
                # has_l3=False 的 run 排除出 (L1-L3) 排名竞争
                acc_by_mm = {}
                for mm in model_modes:
                    d2 = summary.get((mm[0], mm[1], k))
                    if d2 is None:
                        acc_by_mm[mm] = None
                    elif not (d2.get('has_l3') and d2['overall'].get('l3_total', 0) > 0):
                        acc_by_mm[mm] = None  # 没跑 L3，不参与 (L1-L3) 排名
                    else:
                        acc_by_mm[mm] = d2['overall']['acc']
                ranks = rank_models(acc_by_mm)
                my_rank = ranks[(model, mode)] if has_l3 else None
                my_acc = o['acc'] if has_l3 else None

                fill_cell(ws1.cell(row, 4, o['total']), None, mfill, NORMAL, CENTER)
                fill_cell(ws1.cell(row, 5, round(p['out_avg'], 1)), None, mfill, NORMAL, CENTER)
                fill_cell(ws1.cell(row, 6, round(p['speed_avg'], 2)), None, mfill, NORMAL, CENTER)
                fill_cell(ws1.cell(row, 7, round(p['in_avg'], 1)),  None, mfill, NORMAL, CENTER)
                fill_cell(ws1.cell(row, 8, round(p['img_avg'], 1)), None, mfill, NORMAL, CENTER)
                # (L1-L3) 整体 acc — 仅 has_l3 时显示
                cell = ws1.cell(row, 9, f"{my_acc*100:.2f}%" if my_acc is not None else '—')
                cell.fill = mfill
                cell.font = Font(bold=True, size=11, color=acc_color(my_acc) if my_acc else '808080')
                cell.alignment = CENTER
                cell.border = BORDER
                # (L1-L3) 排名 — 仅 has_l3 时显示
                rank_disp = f'#{my_rank}' if my_rank is not None else '—'
                cell_r = ws1.cell(row, 10, rank_disp)
                cell_r.fill = PatternFill("solid", fgColor=RANK_COLOR.get(my_rank, 'FFFFFF')) if my_rank else mfill
                cell_r.font = Font(bold=True, size=11, color='1F1F1F' if my_rank else '808080')
                cell_r.alignment = CENTER
                cell_r.border = BORDER

                # ─── L3 N（仅 has_l3 才有） + L1+L2 准确率（无条件，L3 算入分母按错算） + 排名 ───

                # L3 N（只在有 L3 评测时填）
                if has_l3:
                    fill_cell(ws1.cell(row, 11, l3_total_nli), None, mfill, NORMAL, CENTER)
                else:
                    fill_cell(ws1.cell(row, 11, '—'), None, mfill, NORMAL, CENTER)

                # L1+L2 准确率 = (L1_c + L2_c) / (L1 + L2 + L3_pending)
                # 关键：分母中的 L3 用 l3_pending（题数），不是 l3_total（NLI 覆盖数）
                # 这样无论 NLI 跑没跑，L3 都会算入分母、按错算
                l1_n = o.get('l1', 0)
                l2_n = o.get('l2', 0)
                l1l2_n = l1_n + l2_n + l3_pending
                l1l2_c = o.get('l1_c', 0) + o.get('l2_c', 0)
                l1l2_acc = (l1l2_c / l1l2_n) if l1l2_n else 0
                cell_l1l2 = ws1.cell(row, 12, f"{l1l2_acc*100:.2f}%" if l1l2_n else '—')
                cell_l1l2.fill = mfill
                cell_l1l2.font = Font(bold=True, size=11, color=acc_color(l1l2_acc))
                cell_l1l2.alignment = CENTER
                cell_l1l2.border = BORDER

                # L1+L2 排名：同 k 横向比所有模型 (L1_c+L2_c)/(L1+L2+L3_pending)
                # 所有 run 都参与排名（有/无 L3 都用同一公式）
                l1l2_acc_by_mm = {}
                for mm in model_modes:
                    d2 = summary.get((mm[0], mm[1], k))
                    if d2 is None:
                        l1l2_acc_by_mm[mm] = None
                    else:
                        o2 = d2['overall']
                        n_l1l2 = o2.get('l1', 0) + o2.get('l2', 0) + o2.get('l3_pending', 0)
                        c_l1l2 = o2.get('l1_c', 0) + o2.get('l2_c', 0)
                        l1l2_acc_by_mm[mm] = (c_l1l2 / n_l1l2) if n_l1l2 else None
                l1l2_ranks = rank_models(l1l2_acc_by_mm)
                l1l2_my_rank = l1l2_ranks[(model, mode)]
                l1l2_rank_disp = f'#{l1l2_my_rank}' if l1l2_my_rank is not None else '—'
                cell_l1l2r = ws1.cell(row, 13, l1l2_rank_disp)
                cell_l1l2r.fill = PatternFill("solid", fgColor=RANK_COLOR.get(l1l2_my_rank, 'FFFFFF')) if l1l2_my_rank else mfill
                cell_l1l2r.font = Font(bold=True, size=11, color='1F1F1F' if l1l2_my_rank else '808080')
                cell_l1l2r.alignment = CENTER
                cell_l1l2r.border = BORDER
            row += 1

        # 段尾合并 A/B
        section_end = row - 1
        if section_start <= section_end:
            ws1.merge_cells(start_row=section_start, start_column=1, end_row=section_end, end_column=1)
            a_cell = ws1.cell(section_start, 1)
            a_cell.fill = mfill; a_cell.font = Font(bold=True, size=11)
            a_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            a_cell.border = BORDER
            for r2 in range(section_start + 1, section_end + 1):
                ws1.cell(r2, 1).border = BORDER
            ws1.merge_cells(start_row=section_start, start_column=2, end_row=section_end, end_column=2)
            b_cell = ws1.cell(section_start, 2)
            b_cell.fill = mfill; b_cell.font = Font(size=10)
            b_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            b_cell.border = BORDER
            for r2 in range(section_start + 1, section_end + 1):
                ws1.cell(r2, 2).border = BORDER

    widths1 = [20, 11, 6, 9, 13, 12, 12, 12, 14, 9, 8, 14, 9]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A3'

    # ====================================================
    # Sheet 2: type_acc（转置：行=k，列=type）整体 (L1-L3) 准确率
    # ====================================================
    ws2 = wb.create_sheet('type_acc')
    # 列：模型 | k | 12 type (acc%) | 12 type 排名 | TOTAL | 排名
    nt = len(TYPE_ORDER)
    ncol2 = 2 + nt + nt + 2  # model + k + 12 type + 12 type_rank + TOTAL + rank
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol2)
    fill_cell(ws2['A1'], f'VRS-VQA 12 类型 × (model, k) 准确率矩阵（转置：行=k，列=type）（{l3_note}）',
              TITLE_FILL, BOLD_TITLE, CENTER)
    ws2.row_dimensions[1].height = 28

    # 表头
    fill_cell(ws2.cell(2, 1, '模型'), None, HEADER_FILL, BOLD_WHITE, CENTER)
    fill_cell(ws2.cell(2, 2, 'k'),    None, HEADER_FILL, BOLD_WHITE, CENTER)
    for ti, t in enumerate(TYPE_ORDER):
        fill_cell(ws2.cell(2, 3 + ti, t), t, HEADER_FILL, BOLD_WHITE, CENTER)
    # 12 type 排名列
    for ti, t in enumerate(TYPE_ORDER):
        fill_cell(ws2.cell(2, 3 + nt + ti, f'{t} 排名'), t, HEADER_FILL, BOLD_WHITE, CENTER)
    fill_cell(ws2.cell(2, 3 + nt + nt, 'TOTAL'),  None, HEADER_FILL, BOLD_WHITE, CENTER)
    fill_cell(ws2.cell(2, 4 + nt + nt, '排名'),   None, HEADER_FILL, BOLD_WHITE, CENTER)
    ws2.row_dimensions[2].height = 30

    # 数据
    row = 3
    for model, mode in model_modes:
        mfill = PatternFill("solid", fgColor=MODEL_COLOR.get(model, 'FFFFFF'))
        section_start = row
        for k in k_values:
            d = summary.get((model, mode, k))
            ws2.cell(row, 1, MODEL_DISPLAY.get(model, model))
            fill_cell(ws2.cell(row, 2, k), None, mfill, BOLD_DARK, CENTER)

            # 12 type acc% 列
            for ti, t in enumerate(TYPE_ORDER):
                bt = d['by_type'].get(t) if d else None
                acc_v = bt['acc'] if bt else 0
                cell = ws2.cell(row, 3 + ti, f"{acc_v*100:.2f}%" if bt and bt['total'] else '—')
                cell.fill = mfill
                cell.font = Font(bold=True, size=10, color=acc_color(acc_v))
                cell.alignment = CENTER
                cell.border = BORDER

            # 12 type 排名列：同 k 横向比所有 (model, mode) 在该 type 下的 acc
            for ti, t in enumerate(TYPE_ORDER):
                type_acc_by_mm = {}
                for mm in model_modes:
                    d2 = summary.get((mm[0], mm[1], k))
                    if d2 is None:
                        type_acc_by_mm[mm] = None
                    else:
                        bt2 = d2['by_type'].get(t)
                        type_acc_by_mm[mm] = (bt2['acc'] if (bt2 and bt2.get('total')) else None)
                type_ranks = rank_models(type_acc_by_mm)
                my_tr = type_ranks[(model, mode)]
                tr_disp = f'#{my_tr}' if my_tr is not None else '—'
                cell_tr = ws2.cell(row, 3 + nt + ti, tr_disp)
                cell_tr.fill = PatternFill("solid", fgColor=RANK_COLOR.get(my_tr, 'FFFFFF')) if my_tr else mfill
                cell_tr.font = Font(bold=True, size=10, color='1F1F1F' if my_tr else '808080')
                cell_tr.alignment = CENTER
                cell_tr.border = BORDER

            # TOTAL 列（总 acc）
            o = d['overall'] if d else {'acc': 0, 'total': 0}
            cell = ws2.cell(row, 3 + nt + nt, f"{o['acc']*100:.2f}%" if o['total'] else '—')
            cell.fill = mfill
            cell.font = Font(bold=True, size=11, color=acc_color(o['acc']))
            cell.alignment = CENTER
            cell.border = BORDER

            # 整体 rank（同 k 横向比所有模型 acc）
            acc_by_mm = {}
            for mm in model_modes:
                d2 = summary.get((mm[0], mm[1], k))
                acc_by_mm[mm] = d2['overall']['acc'] if d2 else None
            ranks = rank_models(acc_by_mm)
            my_rank = ranks[(model, mode)]
            rank_disp = f'#{my_rank}' if my_rank is not None else '—'
            cell_r = ws2.cell(row, 4 + nt + nt, rank_disp)
            cell_r.fill = PatternFill("solid", fgColor=RANK_COLOR.get(my_rank, 'FFFFFF')) if my_rank else mfill
            cell_r.font = Font(bold=True, size=11, color='1F1F1F' if my_rank else '808080')
            cell_r.alignment = CENTER
            cell_r.border = BORDER
            row += 1

        # 合并 A 列 model 名（覆盖本段所有 k 行）
        section_end = row - 1
        if section_start <= section_end:
            ws2.merge_cells(start_row=section_start, start_column=1,
                            end_row=section_end, end_column=1)
            c = ws2.cell(section_start, 1, f"{MODEL_DISPLAY.get(model, model)}\n[{mode}]")
            c.fill = mfill; c.font = BOLD_DARK
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = BORDER
            for r2 in range(section_start + 1, section_end + 1):
                ws2.cell(r2, 1).border = BORDER

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 6
    for ti in range(nt):
        ws2.column_dimensions[get_column_letter(3 + ti)].width = 11
    for ti in range(nt):
        ws2.column_dimensions[get_column_letter(3 + nt + ti)].width = 8
    ws2.column_dimensions[get_column_letter(3 + nt + nt)].width = 11
    ws2.column_dimensions[get_column_letter(4 + nt + nt)].width = 7
    ws2.freeze_panes = 'C3'

    # ====================================================
    # Sheet 3: raw_concat
    # ====================================================
    ws3 = wb.create_sheet('raw_concat')
    headers3 = ['model', 'mode', 'k', '_idx', 'type', 'correct',
                'judge_level', 'question', 'gt', 'pred', 'image_id',
                'tokens', 'input_tokens', 'img_tokens', 'speed']
    for ci, h in enumerate(headers3, 1):
        fill_cell(ws3.cell(1, ci, h), h, HEADER_FILL, BOLD_WHITE, CENTER)
    ws3.row_dimensions[1].height = 22
    for ri, r in enumerate(all_raw_rows, 2):
        for ci, h in enumerate(headers3, 1):
            v = r.get(h, '')
            if isinstance(v, str) and len(v) > 200:
                v = v[:200] + '...'
            fill_cell(ws3.cell(ri, ci, v), None, None,
                      Font(size=9, color='1F1F1F' if h in ('model','mode','k','type','correct','judge_level') else '505050'),
                      LEFT)
    widths3 = [18, 10, 6, 8, 16, 8, 12, 50, 12, 20, 16, 8, 12, 10, 8]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A2'

    # ─── 保存 ───────────────────────────────────────────
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"\n[完成] 写入 {out_path}")
    print(f"      涉及: {len(model_modes)} 个 (model, mode) × {len(k_values)} 个 k = {len(summary)} 个 run")
    print(f"      Sheet 1: per_lever  整体 (L1-L3) acc + L1+L2 acc + 双排名")
    print(f"      Sheet 2: type_acc   12 类型 × (model, k) 转置矩阵（acc% + per-type 排名 + 整体 acc/排名）")
    print(f"      Sheet 3: raw_concat 扁平 {len(all_raw_rows)} 条")


# ─── 入口 ──────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='VRS-VQA 分层评测报告生成器 v3',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('--raw-root', type=Path, default=DEFAULT_RAW_ROOT,
                        help=f'raw_outputs 根目录（默认: {DEFAULT_RAW_ROOT}）')
    parser.add_argument('--out-dir', type=Path, default=DEFAULT_OUT_DIR,
                        help=f'xlsx 输出目录（默认: {DEFAULT_OUT_DIR}）')
    parser.add_argument('--name', type=str, default='VQA_type_report_v3',
                        help='xlsx 文件名（不含 .xlsx 后缀，默认: VQA_type_report_v3）')
    parser.add_argument('--include-mode', type=str, nargs='+', default=None,
                        choices=['thinkOFF', 'thinkON'],
                        help='只统计指定模式（默认: 全部）')
    parser.add_argument('--l3-as-wrong', type=str, default='false',
                        choices=['true', 'false'],
                        help='L3 pending 是否按 0 算（默认: false → 整体 (L1-L3) acc）')
    args = parser.parse_args()
    l3_as_wrong = args.l3_as_wrong == 'true'

    out_path = args.out_dir / f'{args.name}.xlsx'
    include_modes = set(args.include_mode) if args.include_mode else None

    print(f"raw_root:    {args.raw_root}")
    print(f"out_dir:     {args.out_dir}")
    print(f"name:        {args.name}")
    print(f"include_mode:{include_modes or 'all'}")
    print(f"l3_as_wrong: {l3_as_wrong}")

    if not args.raw_root.exists():
        print(f"[ERROR] raw_root 不存在: {args.raw_root}")
        sys.exit(1)

    print(f"\n[1/4] 加载 VRSBench type 索引: {VQA_ANNOT}")
    type_map = load_type_map()
    print(f"      共 {len(type_map)} 条 question_id → type 映射")

    print(f"\n[2/4] 扫 raw_outputs 目录")
    runs = discover_runs(args.raw_root, include_modes)
    if not runs:
        print(f"[ERROR] {args.raw_root} 下没找到 raw_outputs.json")
        sys.exit(1)
    print(f"      发现 {len(runs)} 个 run")
    for r in runs:
        print(f"        {r[0]} | {r[1]} | k={r[2]:>3}")

    print(f"\n[3/4] 聚合统计 (L3_as_wrong={l3_as_wrong})")
    summary, all_raw_rows = collect_summary(runs, type_map, l3_as_wrong)
    for (model, mode, k), d in summary.items():
        o = d['overall']; p = d['perf']
        print(f"      {model:25} {mode:9} k={k:>3}: N={o['total']:>5}  "
              f"acc={o['acc']*100:5.1f}%  L1c={o['l1_c']:>4}  L2c={o['l2_c']:>3}  L3_pending={o['l3_pending']:>4}  "
              f"perf: out={p['out_avg']:.1f}  speed={p['speed_avg']:.1f}t/s")

    print(f"\n[4/4] 写 Excel")
    write_excel(summary, all_raw_rows, out_path, l3_as_wrong)


if __name__ == '__main__':
    main()
