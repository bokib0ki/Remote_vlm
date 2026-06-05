#!/usr/bin/env python3
"""
合并 lever_test2/lever_k=400_vqa/ + vqa_3sets/lever_k=400_vqa_{1,2,3}/
生成 k=800/1200/1600 三个新挡位，写入 VQA_type_report_v3.xlsx 后面。

合并规则（按题去重 by _idx）：
  k=800  = vqa_3sets/lever_k=400_vqa_1/ + ..._2/    (杠杆 400×2 = 800)
  k=1200 = vqa_3sets/lever_k=400_vqa_1/ + ..._2/ + ..._3/   (杠杆 400×3 = 1200)
  k=1600 = lever_test2/lever_k=400_vqa/ + vqa_3sets/lever_k=400_vqa_{1,2,3}/  (杠杆 400×4 = 1600)

L3 评测用 vqa_l3_llm.json（Qwen3.5-4B judge, 温度=0）
L1 短路 (substring) + L2 短路 (yes/no/数字) 来自 raw_outputs.json

输出：跟 VQA_type_report_v3.xlsx 同样格式的 per_lever + type_acc 2 sheet
"""
import json
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── 路径 ──────────────────────────────────────────────────
PROJ       = Path('/home/admin1/projects/remote_vlm_eval')
LT2        = Path('/home/admin1/models/raw_outputs/lever_test2')
VQA3       = Path('/home/admin1/models/raw_outputs/vqa_3sets')
VRS        = Path('/home/admin1/models/VRSBench_EVAL_vqa.json')
EXIST_XLSX = PROJ / 'excel_result' / 'VQA_type_report_v3.xlsx'
OUT_XLSX   = PROJ / 'excel_result' / 'VQA_type_report_v3.xlsx'  # 覆盖（追加会太复杂）

# ─── 模型展示名 ────────────────────────────────────────────
MODEL_DISPLAY = {
    'qwen2.5-vl-3B':         'Qwen2.5-VL-3B',
    'qwen2.5-vl-7B':         'Qwen2.5-VL-7B',
    'qwen3.5-0.8B':          'Qwen3.5-0.8B',
    'qwen3.5-2B':            'Qwen3.5-2B',
    'qwen3.5-4B':            'Qwen3.5-4B',
    'qwen3-vl-2B':           'Qwen3-VL-2B',
    'qwen3-vl-4B':           'Qwen3-VL-4B',
    'qwen3-vl-4B-thinking':  'Qwen3-VL-4B-Thinking',
    'minicpm-v-4.6':         'MiniCPM-V 4.6',
    'gemma-4-e2b':           'Gemma-4-2B',
    'gemma-4-e4b':           'Gemma-4-4B',
}
MODEL_COLOR = {
    'qwen2.5-vl-3B':         'D6EAF8',
    'qwen3.5-2B':            'F9E79F',
    'qwen3.5-4B':            'F7DC6F',
    'qwen3-vl-4B':           '76D7C4',
    'qwen3-vl-4B-thinking':  'FAD7A0',
    'minicpm-v-4.6':         'D5F5E3',
    'gemma-4-e4b':           'EDBB99',
}
RANK_COLOR = {1: 'FFD700', 2: 'C0C0C0', 3: 'CD7F32'}

THINKOFF_MODELS = ['gemma-4-e4b', 'minicpm-v-4.6', 'qwen2.5-vl-3B',
                   'qwen3-vl-4B', 'qwen3.5-2B', 'qwen3.5-4B']
TYPES_ORDER = ['object existence', 'object quantity', 'object position',
               'object category', 'object color', 'scene type',
               'object shape', 'image', 'object size',
               'reasoning', 'object direction', 'rural or urban']

# ─── 通用样式 ─────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TITLE_FILL  = PatternFill("solid", fgColor="17202A")
TOTAL_FILL  = PatternFill("solid", fgColor="EAEDED")
BOLD_WHITE  = Font(bold=True, color="FFFFFF", size=10)
BOLD_DARK   = Font(bold=True, color="1F1F1F", size=10)
BOLD_TITLE  = Font(bold=True, color="FFFFFF", size=13)
NORMAL      = Font(size=10)
THIN        = Side(style='thin')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)


def fill_cell(cell, value=None, fill=None, font=None, align=None, border=True):
    if value is not None: cell.value = value
    if fill:  cell.fill = fill
    if font:  cell.font = font
    if align: cell.alignment = align
    if border: cell.border = BORDER


def load_type_map():
    return {str(it['question_id']): it['type'] for it in json.load(open(VRS))}


def parse_dir(name):
    """识别 vqa_3sets_lever_k=400_vqa_N / lever_test2_lever_k=400_vqa 两种 pattern"""
    patterns = [
        r'^(.+?)_(think(?:ON|OFF))_lever_test2_lever_k=(\d+)_vqa$',
        r'^(.+?)_(think(?:ON|OFF))_vqa_3sets_lever_k=(\d+)_vqa_(\d+)$',
    ]
    for pat in patterns:
        m = re.match(pat, name)
        if m:
            if 'vqa_3sets' in pat:
                return m.group(1), m.group(2), int(m.group(3)), int(m.group(4))
            return m.group(1), m.group(2), int(m.group(3)), None
    return None


# ─── 加载 run 数据 ─────────────────────────────────────────
def load_run(run_dir):
    """返回 (raw_recs, l3_idx_to_correct, has_l3, l3_method)"""
    raw_path = run_dir / 'raw_outputs.json'
    if not raw_path.exists():
        return None, None, False, 'none'
    raw_recs = json.load(open(raw_path))

    l3_idx_to_correct = {}
    has_l3 = False
    method = 'none'
    for l3_path, m in [(run_dir / 'vqa_l3_llm.json', 'llm'),
                       (run_dir / 'vqa_l3_nli.json', 'nli'),
                       (run_dir / 'vqa_l3_bert.json', 'bertscore')]:
        if l3_path.exists():
            for lr in json.load(open(l3_path)):
                idx = str(lr.get('_idx', ''))
                c = lr.get('correct_l3', lr.get('llm_correct_l3', 0))
                l3_idx_to_correct[idx] = int(c)
            has_l3 = True
            method = m
            break
    return raw_recs, l3_idx_to_correct, has_l3, method


def merge_runs(run_dirs, has_l3_required=True):
    """
    合并多个 run_dir（按 _idx 去重，保留首次出现）。
    返回合并后的 list of records + has_l3 状态。
    """
    seen = set()
    merged = []
    l3_all = {}
    has_l3_overall = True
    for rd in run_dirs:
        raw, l3, hl, _ = load_run(rd)
        if raw is None:
            continue
        for r in raw:
            idx = str(r.get('_idx', ''))
            if idx in seen:
                continue
            seen.add(idx)
            merged.append(r)
            if hl and idx in l3:
                l3_all[idx] = l3[idx]
        if not hl and has_l3_required:
            has_l3_overall = False
    return merged, l3_all, has_l3_overall


# ─── 6 个 thinkOFF 模型 k=800/1200/1600 找 run 目录 ────────
def find_runs_for_model(model_name, k_400_path):
    """返回 {800: [...], 1200: [...], 1600: [...]} 每档 1 份或 2 份目录列表"""
    # 老 k=400 (1 份)
    old_dir = None
    for p in (k_400_path).glob(f'{model_name}_thinkOFF_lever_test2_lever_k=400_vqa'):
        if p.is_dir():
            old_dir = p
            break

    # 新 1/2/3
    new_dirs = {}
    for n in [1, 2, 3]:
        for p in VQA3.glob(f'lever_k=400_vqa_{n}/{model_name}_thinkOFF_vqa_3sets_lever_k=400_vqa_{n}'):
            if p.is_dir():
                new_dirs[n] = p
                break

    return old_dir, new_dirs


# ─── 计算 overall + by_type 指标 ────────────────────────────
def compute_metrics(merged_recs, l3_idx_to_correct, has_l3, type_map):
    """
    返回 (overall, by_type, perf)
    overall: dict
    by_type: {type_name: dict}
    """
    by_type = defaultdict(lambda: {
        'total': 0, 'l1': 0, 'l1_c': 0, 'l2': 0, 'l2_c': 0,
        'l3_pending': 0, 'l3_correct': 0, 'correct_total': 0,
    })
    overall = {
        'total': 0, 'l1': 0, 'l1_c': 0, 'l2': 0, 'l2_c': 0,
        'l3_pending': 0, 'l3_correct': 0, 'correct_total': 0,
    }
    perf = {'out': [], 'in': [], 'img': [], 'speed': []}

    for r in merged_recs:
        idx = str(r.get('_idx', ''))
        t = type_map.get(idx, '<UNKNOWN>')
        c = r.get('correct', 0)
        is_c = (str(c) in ('1', 'True', 'true')) or (c is True)
        lvl = r.get('judge_level', '')

        overall['total'] += 1
        bt = by_type[t]
        bt['total'] += 1

        if lvl == 'L1':
            overall['l1'] += 1; overall['l1_c'] += int(is_c)
            bt['l1'] += 1;      bt['l1_c'] += int(is_c)
        elif lvl == 'L2':
            overall['l2'] += 1; overall['l2_c'] += int(is_c)
            bt['l2'] += 1;      bt['l2_c'] += int(is_c)
        elif lvl == 'L3':
            overall['l3_pending'] += 1
            bt['l3_pending'] += 1
            if has_l3 and idx in l3_idx_to_correct:
                l3c = l3_idx_to_correct[idx]
                overall['l3_correct'] += l3c
                bt['l3_correct'] += l3c

        perf['out'].append(r.get('tokens', 0))
        perf['in'].append(r.get('input_tokens', 0))
        perf['img'].append(r.get('img_tokens', 0))
        perf['speed'].append(r.get('speed', 0))

    overall['correct_total'] = overall['l1_c'] + overall['l2_c']
    for bt_v in by_type.values():
        bt_v['correct_total'] = bt_v['l1_c'] + bt_v['l2_c']

    n = len(perf['out'])
    perf_summary = {
        'out_avg':  sum(perf['out'])  / n if n else 0,
        'in_avg':   sum(perf['in'])   / n if n else 0,
        'img_avg':  sum(perf['img'])  / n if n else 0,
        'speed_avg': sum(perf['speed']) / n if n else 0,
    }

    return overall, dict(by_type), perf_summary


def acc_color(acc):
    if acc is None: return '808080'
    if acc >= 0.90: return '0B5345'
    if acc >= 0.70: return '1A5276'
    if acc >= 0.50: return 'B9770E'
    return '922B21'


def fmt_pct(c, n):
    return f'{c/n:.2%}' if n > 0 else '—'


def L3_method_text(data_list):
    """返回 L3 评测方法标签（现在全部用 LLM）"""
    return ' | L3 评测: LLM'


def collect_old_data():
    """
    重新读 47 个老 run 的 raw_outputs.json + vqa_l3_llm.json，调 compute_metrics 算指标。
    返回 list of (model, mode, k, has_l3, overall, by_type, perf)，结构跟新 run 的一致。
    """
    type_map = load_type_map()
    data = []
    for run_dir in sorted(LT2.glob('lever_k=*_vqa')):
        for m_dir in sorted(run_dir.iterdir()):
            if not m_dir.is_dir():
                continue
            parsed = parse_dir(m_dir.name)
            if not parsed:
                continue
            model, mode, k, _ = parsed
            merged, l3, has_l3 = merge_runs([m_dir], has_l3_required=True)
            if not merged:
                continue
            o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
            data.append((model, mode, k, has_l3, o, bt, p))
    return data


# ─── 主流程 ───────────────────────────────────────────────
def main():
    type_map = load_type_map()
    k_400_path = LT2 / 'lever_k=400_vqa'

    # 收集 6 个 thinkOFF 模型 × 3 挡位的数据
    all_data = []  # [(model, mode, k, has_l3, overall, by_type, perf)]
    for m in THINKOFF_MODELS:
        old_dir, new_dirs = find_runs_for_model(m, k_400_path)
        if not new_dirs:
            print(f'[SKIP] {m}: 没找到 vqa_3sets 数据')
            continue

        # k=800 = 1+2
        if 1 in new_dirs and 2 in new_dirs:
            merged, l3, has_l3 = merge_runs([new_dirs[1], new_dirs[2]])
            o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
            all_data.append((m, 'thinkOFF', 800, has_l3, o, bt, p))

        # k=1200 = 1+2+3
        if all(n in new_dirs for n in [1, 2, 3]):
            merged, l3, has_l3 = merge_runs([new_dirs[1], new_dirs[2], new_dirs[3]])
            o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
            all_data.append((m, 'thinkOFF', 1200, has_l3, o, bt, p))

        # k=1600 = old + 1+2+3
        if old_dir and all(n in new_dirs for n in [1, 2, 3]):
            merged, l3, has_l3 = merge_runs([old_dir, new_dirs[1], new_dirs[2], new_dirs[3]])
            o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
            all_data.append((m, 'thinkOFF', 1600, has_l3, o, bt, p))

    # 1 个 thinkON 模型 (qwen3-vl-4B-thinking) 只有 k=1600 (老 + 1+2+3)
    m_thinking = 'qwen3-vl-4B-thinking'
    old_dir_t = None
    for p in (LT2 / 'lever_k=400_vqa').glob(f'{m_thinking}_thinkON_lever_test2_lever_k=400_vqa'):
        if p.is_dir():
            old_dir_t = p
            break
    # 没有 thinking 的 vqa_3sets 数据，所以 k=1600 不能给 thinking
    # 如果有，则合并
    new_dirs_t = {}
    for n in [1, 2, 3]:
        for p in VQA3.glob(f'lever_k=400_vqa_{n}/{m_thinking}_thinkON_vqa_3sets_lever_k=400_vqa_{n}'):
            if p.is_dir():
                new_dirs_t[n] = p
                break
    if old_dir_t and all(n in new_dirs_t for n in [1, 2, 3]):
        merged, l3, has_l3 = merge_runs([old_dir_t, new_dirs_t[1], new_dirs_t[2], new_dirs_t[3]])
        o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
        all_data.append((m_thinking, 'thinkON', 1600, has_l3, o, bt, p))
        # thinking 也有 k=800/1200（如果 new_dirs 都有）
        if all(n in new_dirs_t for n in [1, 2]):
            merged, l3, has_l3 = merge_runs([new_dirs_t[1], new_dirs_t[2]])
            o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
            all_data.append((m_thinking, 'thinkON', 800, has_l3, o, bt, p))
        if all(n in new_dirs_t for n in [1, 2, 3]):
            merged, l3, has_l3 = merge_runs([new_dirs_t[1], new_dirs_t[2], new_dirs_t[3]])
            o, bt, p = compute_metrics(merged, l3, has_l3, type_map)
            all_data.append((m_thinking, 'thinkON', 1200, has_l3, o, bt, p))

    print(f'\n=== 合并结果：{len(all_data)} 个 (model, mode, k) ===')
    for m, mode, k, hl, o, _, _ in all_data:
        l1l2_strict = (o['l1_c'] + o['l2_c']) / (o['total']) if o['total'] else 0
        l1l3 = (o['l1_c'] + o['l2_c'] + o['l3_correct']) / o['total'] if o['total'] else 0
        method = 'LLM' if hl else 'none'
        print(f'  {m:25} {mode:9} k={k:>4}  N={o["total"]:>5}  '
              f'L1={o["l1"]:>4} L2={o["l2"]:>4} L3={o["l3_pending"]:>4}  '
              f'L1+L2严={l1l2_strict*100:>5.2f}%  L1-L3={l1l3*100:>5.2f}%  (L3 method: {method})')

    # 写 Excel（追加到现有 VQA_type_report_v3.xlsx 后）
    append_to_xlsx(all_data)


def append_to_xlsx(all_data):
    """
    把 3 个新挡位数据**合并到老数据，按 (model, k) 排序统一重写** per_lever + type_acc sheet。
    """
    from openpyxl import load_workbook
    wb = load_workbook(EXIST_XLSX)

    # 收集老的 47 run 数据
    old_data = collect_old_data()

    # 合并：老的 47 + 新的 18
    all_combined = old_data + all_data
    # 按 (model, k) 排序
    all_combined.sort(key=lambda x: (x[0], x[2]))

    # ──── per_lever sheet 整段重写 ────
    # 先清空 per_lever 的内容（保留 sheet）
    if 'per_lever' in wb.sheetnames:
        del wb['per_lever']
    ws1 = wb.create_sheet('per_lever', 0)  # 放第 1 个位置

    # 计算每个 k 维度的 (L1-L3) 整体 acc → 排名
    acc_by_mk = {}  # (model, k) -> acc
    for m, mode, k, hl, o, _, _ in all_combined:
        if not hl:
            continue
        l1l3 = (o['l1_c'] + o['l2_c'] + o['l3_correct']) / o['total'] if o['total'] else 0
        acc_by_mk[(m, k)] = l1l3

    # 按 k 分组算排名
    rank_by_mk = {}
    for k_val in sorted(set(k for _, _, k, _, _, _, _ in all_combined)):
        sub = {mk: v for mk, v in acc_by_mk.items() if mk[1] == k_val}
        sorted_items = sorted(sub.items(), key=lambda x: x[1], reverse=True)
        rank = {}
        last_v, last_rank = None, 0
        for i, ((mm, kk), v) in enumerate(sorted_items, 1):
            if v != last_v:
                last_rank = i
                last_v = v
            rank[(mm, kk)] = last_rank
        rank_by_mk.update(rank)

    # 写 per_lever
    headers1 = ['模型', '模式', 'k', '总题数 N', '平均输出 token', '平均速度 (t/s)',
                '平均输入 token', '平均图片 token', '准确率 (L1-L3)', '(L1-L3) 排名',
                'L3 N', 'L1+L2 准确率', 'L1+L2 排名']
    ncol1 = len(headers1)
    # 全部用 LLM 评测
    method_note = ' | L3 评测: LLM'

    # title
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol1)
    title_cell = ws1.cell(1, 1)
    title_cell.value = f'VRS-VQA per_lever 总览（(L1+L2+L3) / N_total 整体准确率{method_note}）'
    fill_cell(title_cell, fill=TITLE_FILL, font=BOLD_TITLE, align=CENTER)
    ws1.row_dimensions[1].height = 28
    # header
    for ci, h in enumerate(headers1, 1):
        fill_cell(ws1.cell(2, ci, h), h, HEADER_FILL, BOLD_WHITE, CENTER)
    ws1.row_dimensions[2].height = 30

    # 数据
    row = 3
    last_model = None
    for m, mode, k, hl, o, _, perf in all_combined:
        mfill = PatternFill("solid", fgColor=MODEL_COLOR.get(m, 'FFFFFF'))
        l1l3 = (o['l1_c'] + o['l2_c'] + o['l3_correct']) / o['total'] if o['total'] else 0
        l1l2_strict = (o['l1_c'] + o['l2_c']) / o['total'] if o['total'] else 0
        rank_l1l3 = rank_by_mk.get((m, k), None)
        # 同一 k 下的 L1+L2 排名
        rank_l1l2 = None  # 暂略

        fill_cell(ws1.cell(row, 1, MODEL_DISPLAY.get(m, m) if m != last_model else ''), None, mfill, BOLD_DARK, CENTER)
        fill_cell(ws1.cell(row, 2, mode if m != last_model else ''), None, mfill, NORMAL, CENTER)
        fill_cell(ws1.cell(row, 3, k), None, mfill, BOLD_DARK, CENTER)
        fill_cell(ws1.cell(row, 4, o['total']), None, mfill, NORMAL, CENTER)
        fill_cell(ws1.cell(row, 5, round(perf['out_avg'], 1)), None, mfill, NORMAL, CENTER)
        fill_cell(ws1.cell(row, 6, round(perf['speed_avg'], 2)), None, mfill, NORMAL, CENTER)
        fill_cell(ws1.cell(row, 7, round(perf['in_avg'], 1)),  None, mfill, NORMAL, CENTER)
        fill_cell(ws1.cell(row, 8, round(perf['img_avg'], 1)), None, mfill, NORMAL, CENTER)
        acc_fill = PatternFill("solid", fgColor=acc_color(l1l3))
        fill_cell(ws1.cell(row, 9, f'{l1l3*100:.2f}%'), None, acc_fill, BOLD_WHITE, CENTER)
        fill_cell(ws1.cell(row, 10, f'#{rank_l1l3}' if rank_l1l3 else '—'),
                  None, mfill, NORMAL, CENTER)
        fill_cell(ws1.cell(row, 11, o['l3_pending']), None, mfill, NORMAL, CENTER)
        l1l2_fill = PatternFill("solid", fgColor=acc_color(l1l2_strict))
        fill_cell(ws1.cell(row, 12, f'{l1l2_strict*100:.2f}%'), None, l1l2_fill, BOLD_WHITE, CENTER)
        fill_cell(ws1.cell(row, 13, f'#{rank_l1l2}' if rank_l1l2 else '—'),
                  None, mfill, NORMAL, CENTER)
        last_model = m
        row += 1

    # ──── type_acc sheet 整段重写 ────
    if 'type_acc' in wb.sheetnames:
        del wb['type_acc']
    ws2 = wb.create_sheet('type_acc', 1)  # 放第 2 个位置

    # 12 类型列
    headers2 = ['模型', 'k'] + TYPES_ORDER + [f'{t}.排名' for t in TYPES_ORDER] + ['TOTAL', '排名']
    ncol2 = len(headers2)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol2)
    tcell = ws2.cell(1, 1)
    tcell.value = f'VRS-VQA type_acc 总览（12 类型准确率{L3_method_text(all_combined)}）'
    fill_cell(tcell, fill=TITLE_FILL, font=BOLD_TITLE, align=CENTER)
    ws2.row_dimensions[1].height = 28
    for ci, h in enumerate(headers2, 1):
        fill_cell(ws2.cell(2, ci, h), h, HEADER_FILL, BOLD_WHITE, CENTER)
    ws2.row_dimensions[2].height = 30

    # 算每个 (k, type) 的 acc 排名
    acc_by_kt = {}  # (k, type) -> {(model, k): acc}
    for m, mode, k, hl, o, by_type, _ in all_combined:
        for tname in TYPES_ORDER:
            bt = by_type.get(tname, None)
            if not bt or bt['total'] == 0:
                continue
            acc = (bt['l1_c'] + bt['l2_c'] + bt['l3_correct']) / bt['total']
            acc_by_kt.setdefault((k, tname), []).append(((m, k), acc))
    # 算排名
    rank_by_mkt = {}
    for (k, t), lst in acc_by_kt.items():
        sorted_items = sorted(lst, key=lambda x: x[1], reverse=True)
        rank = {}
        last_v, last_rank = None, 0
        for i, (mk, v) in enumerate(sorted_items, 1):
            if v != last_v:
                last_rank = i
                last_v = v
            rank[mk] = last_rank
        for mk, r in rank.items():
            rank_by_mkt[(*mk, t)] = r
    # 算 total 排名
    total_by_k = {}  # k -> [(model, k, acc)]
    for m, mode, k, hl, o, _, _ in all_combined:
        if not hl:
            continue
        l1l3 = (o['l1_c'] + o['l2_c'] + o['l3_correct']) / o['total'] if o['total'] else 0
        total_by_k.setdefault(k, []).append(((m, k), l1l3))
    total_rank = {}
    for k, lst in total_by_k.items():
        sorted_items = sorted(lst, key=lambda x: x[1], reverse=True)
        rank = {}
        last_v, last_rank = None, 0
        for i, (mk, v) in enumerate(sorted_items, 1):
            if v != last_v:
                last_rank = i
                last_v = v
            rank[mk] = last_rank
        for mk, r in rank.items():
            total_rank[mk] = r

    # 数据
    row2 = 3
    last_model = None
    for m, mode, k, hl, o, by_type, _ in all_combined:
        mfill = PatternFill("solid", fgColor=MODEL_COLOR.get(m, 'FFFFFF'))
        fill_cell(ws2.cell(row2, 1, MODEL_DISPLAY.get(m, m) if m != last_model else ''), None, mfill, BOLD_DARK, CENTER)
        fill_cell(ws2.cell(row2, 2, f'k={k}' if m != last_model else ''), None, mfill, NORMAL, CENTER)
        # 12 类型 acc
        for ti, tname in enumerate(TYPES_ORDER, start=3):
            bt = by_type.get(tname, None)
            if not bt or bt['total'] == 0:
                fill_cell(ws2.cell(row2, ti, '—'), None, mfill, NORMAL, CENTER)
            else:
                acc = (bt['l1_c'] + bt['l2_c'] + bt['l3_correct']) / bt['total']
                tfill = PatternFill("solid", fgColor=acc_color(acc))
                fill_cell(ws2.cell(row2, ti, f'{acc*100:.2f}%'), None, tfill, BOLD_WHITE, CENTER)
        # 12 类型 排名
        for ti, tname in enumerate(TYPES_ORDER, start=15):
            rk = rank_by_mkt.get((m, k, tname), None)
            fill_cell(ws2.cell(row2, ti, f'#{rk}' if rk else '—'), None, mfill, NORMAL, CENTER)
        # TOTAL
        tot_acc = (o['l1_c'] + o['l2_c'] + o['l3_correct']) / o['total'] if o['total'] else 0
        tot_fill = PatternFill("solid", fgColor=acc_color(tot_acc))
        fill_cell(ws2.cell(row2, 27, f'{tot_acc*100:.2f}%'), None, tot_fill, BOLD_WHITE, CENTER)
        rk = total_rank.get((m, k), None)
        fill_cell(ws2.cell(row2, 28, f'#{rk}' if rk else '—'), None, mfill, NORMAL, CENTER)
        last_model = m
        row2 += 1

    wb.save(EXIST_XLSX)
    print(f'\n✓ 重写 per_lever + type_acc，共 {len(all_combined)} 行（老 {len(old_data)} + 新 {len(all_data)}）')


if __name__ == '__main__':
    main()
