"""
NLI 审核表格生成器
读所有 vqa_l3_nli.json，按 correct_l3 排序（1 在前 0 在后），生成审核表
"""
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

RAW_ROOT = Path('/home/admin1/models/raw_outputs/lever_test2')
OUT_PATH = Path('/home/admin1/projects/remote_vlm_eval/excel_result/VQA_NLI_Audit.xlsx')


def parse_dir_name(dirname: str):
    for pat in [
        r'^(.+?)_(think(?:ON|OFF))_lever_test2_lever_k=(\d+)_vqa$',
        r'^(.+?)_(think(?:ON|OFF))_lever_k=(\d+)$',
    ]:
        m = re.match(pat, dirname)
        if m:
            return m.group(1), m.group(2), int(m.group(3))
    return None


def main():
    all_records = []
    n_files = 0
    for l3f in sorted(RAW_ROOT.glob('**/vqa_l3_nli.json')):
        parsed = parse_dir_name(l3f.parent.name)
        if not parsed:
            print(f'[WARN] 跳过（目录名不匹配）: {l3f.parent.name}')
            continue
        model, mode, k = parsed
        n_files += 1
        try:
            recs = json.load(open(l3f))
        except Exception as e:
            print(f'[ERR] {l3f}: {e}')
            continue
        for r in recs:
            all_records.append({
                'model':     model,
                'mode':      mode,
                'k':         k,
                '_idx':      r.get('_idx', ''),
                'image_id':  r.get('image_id', ''),
                'type':      r.get('type', ''),  # 可能没存
                'question':  r.get('question', ''),
                'gt':        r.get('gt', ''),
                'pred':      r.get('pred', ''),
                'e_p':       r.get('nli_entailment', 0),
                'n_p':       r.get('nli_neutral', 0),
                'c_p':       r.get('nli_contradiction', 0),
                'correct_l3': int(r.get('correct_l3', 0)),
            })

    # type 字段要从 raw 里补（vqa_l3_nli.json 不存 type）
    type_map = {}
    ann = json.load(open('/home/admin1/models/VRSBench_EVAL_vqa.json'))
    if isinstance(ann, list):
        for item in ann:
            if 'question_id' in item:
                type_map[item['question_id']] = item.get('type', '<UNKNOWN>')
            elif '_idx' in item:
                type_map[item['_idx']] = item.get('type', '<UNKNOWN>')
    elif isinstance(ann, dict):
        for k_id, v in ann.items():
            if isinstance(v, dict):
                type_map[k_id] = v.get('type', '<UNKNOWN>')
            else:
                type_map[k_id] = v
    for r in all_records:
        if not r['type'] or r['type'] == '<UNKNOWN>':
            r['type'] = type_map.get(r['_idx'], '<UNKNOWN>')

    # 排序：correct_l3 降序（1 在前），e_p 降序
    all_records.sort(key=lambda r: (-r['correct_l3'], -r['e_p']))

    n_total = len(all_records)
    n_correct = sum(1 for r in all_records if r['correct_l3'] == 1)
    n_wrong = n_total - n_correct
    print(f'扫描: {n_files} 个 vqa_l3_nli.json')
    print(f'记录: {n_total} 条 (correct={n_correct}, wrong={n_wrong}, acc={n_correct/n_total*100:.2f}%)')

    # ─── 写 Excel ─────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'NLI_Audit'

    HEADER_FILL = PatternFill("solid", fgColor='305496')
    BOLD_WHITE  = Font(bold=True, color='FFFFFF', size=11)
    BORDER      = Border(*[Side(style='thin', color='BFBFBF')] * 4)
    CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT        = Alignment(horizontal='left', vertical='center', wrap_text=True)

    CORRECT_FILL = PatternFill("solid", fgColor='C6EFCE')  # 浅绿
    WRONG_FILL   = PatternFill("solid", fgColor='FFC7CE')  # 浅红

    headers = [
        'judge', 'correct', 'model', 'mode', 'k', '_idx', 'image_id', 'type',
        'e_p', 'n_p', 'c_p',
        'question', 'gt', 'pred'
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = HEADER_FILL
        c.font = BOLD_WHITE
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 28

    for ri, r in enumerate(all_records, 2):
        is_c = r['correct_l3'] == 1
        # judge: 简化的"模型判"列（显示 ✓/✗）
        judge_disp = '✓ 判对' if is_c else '✗ 判错'
        fill = CORRECT_FILL if is_c else WRONG_FILL

        cells = [
            (judge_disp,        CENTER),
            (r['correct_l3'],   CENTER),
            (r['model'],        LEFT),
            (r['mode'],         CENTER),
            (r['k'],            CENTER),
            (r['_idx'],         CENTER),
            (r['image_id'],     LEFT),
            (r['type'],         CENTER),
            (round(r['e_p'], 4), CENTER),
            (round(r['n_p'], 4), CENTER),
            (round(r['c_p'], 4), CENTER),
            (r['question'],     LEFT),
            (r['gt'],           LEFT),
            (r['pred'],         LEFT),
        ]
        for ci, (v, align) in enumerate(cells, 1):
            c = ws.cell(ri, ci, v)
            c.fill = fill if ci <= 2 else PatternFill("solid", fgColor='FFFFFF' if ri % 2 == 0 else 'F2F2F2')
            c.font = Font(bold=(ci <= 2), size=10, color='1F1F1F')
            c.alignment = align
            c.border = BORDER

    # 列宽
    widths = [10, 8, 22, 10, 6, 8, 18, 22, 8, 8, 8, 50, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'

    # 顶部加统计行
    ws.insert_rows(1)
    stat = f'共 {n_total} 条 | ✓ 判对 {n_correct} ({n_correct/n_total*100:.2f}%) | ✗ 判错 {n_wrong} ({n_wrong/n_total*100:.2f}%)'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(1, 1, stat)
    c.fill = HEADER_FILL
    c.font = Font(bold=True, color='FFFFFF', size=12)
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT_PATH))
    print(f'\n[完成] 写入 {OUT_PATH}')
    print(f'      Sheet: NLI_Audit ({n_total+2} 行 × {len(headers)} 列)')

    # 按 type + correct 统计
    print('\n=== 按 type 统计 ===')
    type_stat = {}
    for r in all_records:
        t = r['type']
        if t not in type_stat:
            type_stat[t] = {'n': 0, 'c': 0}
        type_stat[t]['n'] += 1
        type_stat[t]['c'] += r['correct_l3']
    for t, s in sorted(type_stat.items(), key=lambda x: -x[1]['n']):
        acc = s['c']/s['n']*100
        print(f'  {t:25}  N={s["n"]:>5}  correct={s["c"]:>4}  acc={acc:>5.2f}%')


if __name__ == '__main__':
    main()
