"""
生成 selection 评测的 Excel 报告 — 从 results/{selectionName}/*_metrics.json 读取分数，汇总为表格。

新格式列：模型 | thinkon/off | max_token | 任务1..N | 推理速度
任务列：分数（OFF）或 平均输出token数（ON，仅采样展示）

用法:
  python3 gen_excel_sel.py --selection batch3_sel
  python3 gen_excel_sel.py --selection batch3_sel --out batch3_sel.xlsx
"""
import json, argparse
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError as e:
    raise SystemExit("Missing dependency: openpyxl. Install via: pip install openpyxl") from e

from config import (
    METRICS, MLABELS, MINFO,
    RES_DIR, RAW_DIR_ROOT, OUT_DIR,
)

HDR_FONT = Font(bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
BEST_FILL = PatternFill('solid', fgColor='C6EFCE')
SOTA_FILL = PatternFill('solid', fgColor='FFF2CC')
GREY_FILL = PatternFill('solid', fgColor='F2F2F2')
BORDER = Border(
    left=Side('thin'), right=Side('thin'),
    top=Side('thin'), bottom=Side('thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _fmt(v):
    if v is None:
        return ''
    if isinstance(v, float):
        return round(v, 4)
    return v


def _speed_fmt(v):
    """推理速度保留5位小数，方便比较"""
    if v is None:
        return ''
    return round(v, 5)


def gs(benchmarks, key, sub='acc'):
    """从 benchmarks 字典中取指定 key/sub 的值"""
    mapping = {
        'mme_rs': ('mme_rs', 'acc'),
        'xlrs': ('xlrs', 'acc'),
        'vrs_cap_bleu4': ('vrs_caption', 'bleu4'),
        'vrs_cap_rouge_l': ('vrs_caption', 'rouge_l'),
        'vrs_cap_cider': ('vrs_caption', 'cider'),
        'levir_bleu4': ('levir_cc', 'bleu4'),
        'levir_rouge_l': ('levir_cc', 'rouge_l'),
        'levir_cider': ('levir_cc', 'cider'),
        'vrs_vqa': ('vrs_vqa', 'acc_l1'),
    }
    bk, sk = mapping.get(key, (key, sub))
    return benchmarks.get(bk, {}).get(sk)


# raw_outputs benchmark 名称 → METRICS key 的映射（用于从 raw_outputs 统计 token 数）
RAW_BMARK_MAP = {
    'MME-RS':      'mme_rs',
    'XLRS':        'xlrs',
    'VRS-VQA':     'vrs_vqa',
    'VRS-Caption': 'vrs_cap_bleu4',   # caption 取任一指标 key 即可（都用同一来源）
    'LEVIR-CC':    'levir_bleu4',
}


def parse_tag(stem: str):
    """解析 stem: model_mode_selection"""
    if '_thinkON_' in stem:
        model, rest = stem.split('_thinkON_', 1)
        return model, 'thinkON', rest
    if '_thinkOFF_' in stem:
        model, rest = stem.split('_thinkOFF_', 1)
        return model, 'thinkOFF', rest
    return None


def load_metrics_files(selection_name: str):
    """返回 [(model, mode, scores_dict, stem, max_new, perf_info, raw_token_stats)]"""
    sel_dir = RES_DIR / selection_name
    if not sel_dir.exists():
        raise SystemExit(f"Selection results not found: {sel_dir}")

    rows = []
    for fp in sorted(sel_dir.glob('*_metrics.json')):
        stem = fp.name[:-len('_metrics.json')]
        parsed = parse_tag(stem)
        if not parsed:
            continue
        model, mode, sel = parsed

        d = json.load(open(fp))
        scores = {x: gs(d.get('benchmarks', {}), x) for x in METRICS}
        max_new = d.get('max_new', None)

        # 收集 perf 信息：每个 benchmark 的 avg_s_per_token / avg_tokens
        benches = d.get('benchmarks', {})
        perf_info = {}
        for bk, bv in benches.items():
            if 'perf' in bv:
                p = bv['perf']
                perf_info[bk] = {
                    'avg_s_per_token': p.get('avg_s_per_token'),
                    'avg_tokens': p.get('avg_tokens'),
                }
            if 'token_stats' in bv:
                perf_info[bk] = perf_info.get(bk, {})
                perf_info[bk].update(bv['token_stats'])

        # 从 raw_outputs.json 读取 token 均值（兜底或补充 perf_info）
        raw_token_stats = load_raw_token_stats(selection_name, stem)

        rows.append((model, mode, scores, stem, max_new, perf_info, raw_token_stats))

    rows.sort(key=lambda x: (x[0], 0 if x[1] == 'thinkOFF' else 1))
    return rows


def compute_avg_speed(perf_info):
    """计算所有 benchmark 的加权平均 s_per_token"""
    speeds = []
    for bk, info in perf_info.items():
        sp = info.get('avg_s_per_token')
        if sp is not None:
            speeds.append(sp)
    if not speeds:
        return None
    return sum(speeds) / len(speeds)


def load_raw_token_stats(selection_name: str, tag: str):
    """
    从 raw_outputs.json 按 benchmark 统计平均 token 数。
    返回 {benchmark: avg_tokens}
    """
    rp = RAW_DIR_ROOT / selection_name / tag / 'raw_outputs.json'
    if not rp.exists():
        return {}
    from collections import defaultdict
    bmark_tokens = defaultdict(list)
    for rec in json.load(open(rp)):
        bm = rec.get('benchmark', '?')
        tok = rec.get('tokens')
        if tok is not None:
            bmark_tokens[bm].append(tok)
    return {bm: sum(toks)/len(toks) for bm, toks in bmark_tokens.items()}


def write_sheet1(wb, met_rows, selection_name):
    """
    主表格式：
    列 = [模型, thinkon/off, max_token,
          MME-RS分, MME-RS输出token均,
          XLRS分, XLRS输出token均,
          VRS-Cap B4, VRS-Cap R-L, VRS-Cap CIDER,
          LEVIR B4, LEVIR R-L, LEVIR CIDER,
          VRS-VQA分, VRS-VQA输出token均,
          推理速度 token/s]

    ON模型：分数列留空，输出token数列均值（从perf_info）
    OFF模型：分数列正常，输出token数列留空
    """
    ws = wb.active
    ws.title = '指标表'

    # 定义 benchmark → 列布局
    # (benchmark_key, label, has_score, has_tokens, is_on_avg_only)
    # has_score: OFF 显示分数
    # has_tokens: ON 显示输出token均值（来自 perf_info 或 token_stats）
    bench_layout = [
        ('mme_rs',       'MME-RS',        True,  True),
        ('xlrs',         'XLRS',          True,  True),
        ('vrs_cap_bleu4','VRS-Cap B4',    True,  False),
        ('vrs_cap_rouge_l','VRS-Cap R-L', True,  False),
        ('vrs_cap_cider','VRS-Cap CIDER', True,  False),
        ('levir_bleu4',  'LEVIR B4',      True,  False),
        ('levir_rouge_l','LEVIR R-L',     True,  False),
        ('levir_cider', 'LEVIR CIDER',    True,  False),
        ('vrs_vqa',     'VRS-VQA',        True,  True),
    ]

    # 固定列
    fixed_headers = ['模型', 'thinkon/off', 'max_token']
    # 动态列（每个 benchmark 可能产生1-2列）
    col_headers = []
    col_meta = []  # (bench_key, col_type) where col_type in ('score', 'tokens')
    for bk, label, has_score, has_tokens in bench_layout:
        if has_score:
            col_headers.append(label)
            col_meta.append((bk, 'score'))
        if has_tokens:
            col_headers.append(f'{label}\n输出token均')
            col_meta.append((bk, 'tokens'))
    col_headers.append('推理速度\n(token/s)')
    col_meta.append(('__speed__', 'speed'))

    # 全量 header
    headers = fixed_headers + col_headers

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    # 数据行
    for ri, (model, mode, scores, stem, max_new, perf_info, raw_token_stats) in enumerate(met_rows, 2):
        is_on = (mode == 'thinkON')
        row_fill = PatternFill('solid', fgColor='EBF3FB') if is_on else None

        # 固定列
        c1 = ws.cell(ri, 1, model)
        c1.border = BORDER
        if row_fill:
            c1.fill = row_fill

        c2 = ws.cell(ri, 2, mode)
        c2.border = BORDER
        if row_fill:
            c2.fill = row_fill

        c3 = ws.cell(ri, 3, max_new if max_new is not None else '')
        c3.border = BORDER
        if row_fill:
            c3.fill = row_fill

        # 动态列
        col_idx = 4
        for bk, col_type in col_meta:
            if col_type == 'speed':
                sp = compute_avg_speed(perf_info)
                cell = ws.cell(ri, col_idx, _speed_fmt(sp))
                cell.border = BORDER
                if row_fill:
                    cell.fill = row_fill
                col_idx += 1
                continue

            bv_scores = scores.get(bk)
            bv_perf = perf_info.get(bk, {})
            # 从 raw_outputs 统计的 token 均值（OFF 模型的主来源，ON 模型的兜底）
            raw_key = RAW_BMARK_MAP.get(bk.upper(), bk) if bk.upper() in RAW_BMARK_MAP else bk
            # 尝试找 raw_outputs 里对应的 benchmark 名
            raw_avg = None
            for raw_bench, met_key in RAW_BMARK_MAP.items():
                if met_key == bk:
                    raw_avg = raw_token_stats.get(raw_bench)
                    break

            if col_type == 'score':
                val = bv_scores
                cell = ws.cell(ri, col_idx, _fmt(val))
                cell.border = BORDER
                if row_fill:
                    cell.fill = row_fill
                col_idx += 1
            elif col_type == 'tokens':
                # 优先用 perf_info 里的 avg_tokens（ON 模型有），没有则用 raw_outputs 统计（OFF 模型）
                val = bv_perf.get('avg_tokens')
                if val is None and raw_avg is not None:
                    val = raw_avg
                cell = ws.cell(ri, col_idx, _fmt(val) if val is not None else '')
                cell.border = BORDER
                if row_fill:
                    cell.fill = row_fill
                col_idx += 1

    # 列宽
    ws.column_dimensions['A'].width = 18  # 模型
    ws.column_dimensions['B'].width = 12  # mode
    ws.column_dimensions['C'].width = 10  # max_token
    for ci in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # 冻结前3列
    ws.freeze_panes = 'D2'


def write_sheet2_raw(wb, met_rows, selection_name):
    """原始输出表 - 保持原有结构供参考"""
    ws2 = wb.create_sheet('原始输出表')
    raw_headers = ['模型', '模式', 'Benchmark', 'idx', 'GT', 'Prediction', 'Token数']
    for ci, h in enumerate(raw_headers, 1):
        cell = ws2.cell(1, ci, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER

    r = 2
    for model, mode, _, tag, _, _, _ in met_rows:
        rp = RAW_DIR_ROOT / selection_name / tag / 'raw_outputs.json'
        if not rp.exists():
            continue
        for rec in json.load(open(rp)):
            vals = [
                model, mode,
                rec.get('benchmark', ''),
                rec.get('_idx', ''),
                str(rec.get('gt', ''))[:120],
                str(rec.get('pred', ''))[:120],
                rec.get('tokens', ''),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(r, ci, v)
                cell.border = BORDER
            r += 1

    for ci in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 20


def write_sheet3_summary(wb, met_rows):
    """分析总结表"""
    ws3 = wb.create_sheet('分析总结')
    sum_headers = ['指标', '说明', '最佳模型/模式', '最佳得分']
    for ci, h in enumerate(sum_headers, 1):
        cell = ws3.cell(1, ci, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER

    for ri, m in enumerate(METRICS, 2):
        desc = MINFO.get(m, m)
        best_val, best_mod = None, ''
        for model, mode, scores, _, _, _, _ in met_rows:
            val = scores.get(m)
            if val is not None and (best_val is None or val > best_val):
                best_val = val
                best_mod = f'{model} ({mode})'
        ws3.cell(ri, 1, desc).border = BORDER
        ws3.cell(ri, 2, m).border = BORDER
        cm = ws3.cell(ri, 3, best_mod)
        cm.border = BORDER
        cm.fill = BEST_FILL
        ws3.cell(ri, 4, _fmt(best_val)).border = BORDER

    for ci in range(1, ws3.max_column + 1):
        ws3.column_dimensions[get_column_letter(ci)].width = 22


def make_excel(out_path: Path, selection_name: str, met_rows, include_raw: bool = True):
    wb = openpyxl.Workbook()
    write_sheet1(wb, met_rows, selection_name)
    if include_raw:
        write_sheet2_raw(wb, met_rows, selection_name)
    write_sheet3_summary(wb, met_rows)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--selection', required=True, help='selectionName（results/{selectionName} 文件夹名）')
    parser.add_argument('--out', default='', help='输出文件名（默认 projects/{selectionName}.xlsx）')
    parser.add_argument('--no_raw', action='store_true', help='不写原始输出表')
    args = parser.parse_args()

    selection_name = args.selection
    met_rows = load_metrics_files(selection_name)
    if not met_rows:
        raise SystemExit(f"No metrics found under: {RES_DIR / selection_name}")

    print(f"Loaded {len(met_rows)} model results:")
    for model, mode, _, stem, max_new, perf_info, raw_ts in met_rows:
        print(f"  {stem}: {mode} max_new={max_new} perf_benchmarks={list(perf_info.keys())}")

    out_name = args.out or f'{selection_name}.xlsx'
    make_excel(OUT_DIR / out_name, selection_name, met_rows, include_raw=not args.no_raw)