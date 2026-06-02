#!/usr/bin/env python3
"""
VRS-VQA 分层评测报表生成器 v2
===========================
横轴 = 每个 (模型,k) 组合，k 值从左到右（k=50→100→200→400→500），
每个 k 下 Qwen 和 MiniCPM 并排显示。

列: 模型 | k值 | 总准确率(L1+L2) | 平均输出token | 平均输入token |
     平均图片token | 截断率 | (12种任务类型准确率)
行: 12种任务类型 + 合计行

用法:
  python3 gen_vqa_type_report_v2.py
"""
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 配置
# =====================================================================
PROJ       = Path('/home/admin1/projects/remote_vlm_eval')
MODEL_ROOT = Path('/home/admin1/models')
VQA_ANNOT  = MODEL_ROOT / 'VRSBench_EVAL_vqa.json'

MODELS    = ['qwen3.5-4B', 'minicpm-v-4.6']
K_VALUES  = [50, 100, 200, 400, 500]
MAX_NEW_DEFAULT = 4096   # 截断阈值（token >= 此值视为截断）

# 颜色
DARK_BLUE   = PatternFill("solid", fgColor="1F4E79")
DARK_GREEN  = PatternFill("solid", fgColor="375623")
DARK_ORANGE = PatternFill("solid", fgColor="7D3C0B")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GREY  = PatternFill("solid", fgColor="F2F2F2")
LIGHT_BLUE  = PatternFill("solid", fgColor="DAEEF3")
LIGHT_GREEN = PatternFill("solid", fgColor="D8F0D8")
LIGHT_ORANGE= PatternFill("solid", fgColor="FBE5D6")
Qwen_HEAD   = PatternFill("solid", fgColor="2E86AB")   # 蓝
MiniCPM_HEAD= PatternFill("solid", fgColor="28A745")   # 绿
PERF_HEAD   = PatternFill("solid", fgColor="5C2D91")   # 紫

BOLD_WHITE  = Font(bold=True, color="FFFFFF", size=10)
BOLD_WHITE_LARGE = Font(bold=True, color="FFFFFF", size=12)
BOLD_DARK   = Font(bold=True, color="1F1F1F", size=10)
NORMAL      = Font(size=10)

THIN   = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=False)
RIGHT  = Alignment(horizontal='right',  vertical='center')

# 任务类型顺序（与 annotation 的 type 字段一致）
TYPE_ORDER = [
    'object quantity',
    'object existence',
    'object position',
    'object category',
    'object color',
    'object shape',
    'object size',
    'object direction',
    'scene type',
    'rural or urban',
    'reasoning',
    'image',
]
CHINESE = {
    'object quantity':   '物体计数',
    'object existence':  '物体是否存在',
    'object position':  '物体位置',
    'object category':   '物体类别',
    'object color':      '物体颜色',
    'object shape':      '物体形状',
    'object size':      '物体大小',
    'object direction':  '物体方向',
    'scene type':        '场景类型',
    'rural or urban':    '城乡分类',
    'reasoning':         '推理',
    'image':             '图像属性',
    '__total__':         '合计',
}

# =====================================================================
# 数据加载
# =====================================================================
def load_type_map():
    with open(VQA_ANNOT) as f:
        return {r['question_id']: r['type'] for r in json.load(f)}


def load_records(k, model):
    path = MODEL_ROOT / 'raw_outputs' / f'lever_k={k}' / f'{model}_thinkOFF_lever_k={k}' / 'vqa_raw.json'
    return json.load(open(path)) if path.exists() else None


def compute_row_stats(records, type_map, trunc_thresh=4096):
    """返回 {type_str: (correct, total)} 以及 overall + perf stats"""
    by_type  = defaultdict(lambda: {'correct': 0, 'total': 0})
    overall  = {'correct': 0, 'total': 0}
    perf     = {'input_tokens': [], 'img_tokens': [], 'prompt_tokens': [],
                'output_tokens': [], 'truncated': 0, 'total': 0}

    for r in records:
        lvl = r.get('judge_level', '')
        if lvl not in ('L1', 'L2'):
            continue

        qtype  = type_map.get(r['_idx'], 'unknown')
        c      = r.get('correct', 0)

        overall['total']    += 1
        overall['correct']  += c
        by_type[qtype]['total']    += 1
        by_type[qtype]['correct']  += c

        perf['input_tokens'].append(r['input_tokens'])
        perf['img_tokens'].append(r['img_tokens'])
        perf['prompt_tokens'].append(r['prompt_tokens'])
        out_tok = r.get('tokens', 0)
        perf['output_tokens'].append(out_tok)
        if out_tok >= trunc_thresh:
            perf['truncated'] += 1
        perf['total'] += 1

    return by_type, overall, perf


# =====================================================================
# Excel 构建
# =====================================================================
def acc_str(c, n):
    return f'{c/n:.1%}' if n > 0 else 'N/A'

def num_str(vals, decimals=1):
    if not vals: return 'N/A'
    return f'{sum(vals)/len(vals):.{decimals}f}'

def pct_str(n, total):
    return f'{n/total:.2%}' if total > 0 else 'N/A'

def style(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = fill
    if font:  c.font   = font
    if align: c.alignment = align
    c.border = BORDER
    return c


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ----------------------------------------------------------------------
# 主函数
# ----------------------------------------------------------------------
def main():
    print('加载 type 映射…')
    type_map = load_type_map()

    print('加载评测记录…')
    all_data = {}  # (k, model) -> {by_type, overall, perf}
    for k in K_VALUES:
        for model in MODELS:
            recs = load_records(k, model)
            if recs is None:
                continue
            by_type, overall, perf = compute_row_stats(recs, type_map)
            all_data[(k, model)] = {'by_type': by_type, 'overall': overall, 'perf': perf}

    print(f'已加载 {len(all_data)} 个 (k,model) 组合')

    # ----------------------------------------------------------
    # 构建 Excel
    # ----------------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('VRS-VQA任务类型准确率')

    # ---- 列布局 ----
    # col 1: type_en  (20)
    # col 2: type_cn  (16)
    # col 3-4:   k=50  (Qwen, MiniCPM)   → 2 cols per model per k
    # col 5-6:   k=100 (Qwen, MiniCPM)
    # col 7-8:   k=200 (Qwen, MiniCPM)
    # col 9-10:  k=400 (Qwen, MiniCPM)
    # col 11-12: k=500 (Qwen, MiniCPM)
    # Total: 12 cols
    #
    # 横轴顺序（每2列=一个k值的Qwen+MiniCPM）:
    # 模型 | k | 总准确率 | 平均输出tok | 平均输入tok | 平均图tok | 截断率 |
    #   | qty | exist | pos | cat | color | shape | size | dir | scene | rural | reason | image
    #
    # 实际上：每个 (k, model) = 1列（准确率数字），
    # 列1=type_en, 列2=type_cn, 列3=50_Qwen, 列4=50_MiniCPM,
    # 列5=100_Qwen, 列6=100_MiniCPM, ..., 列11=500_Qwen, 列12=500_MiniCPM

    # ---- 标题行 ----
    ws.merge_cells('A1:L1')
    title = ws['A1']
    title.value = 'VRS-VQA 任务类型准确率 — Qwen3.5-4B vs MiniCPM-V 4.6（think-off）'
    title.fill   = DARK_BLUE
    title.font   = BOLD_WHITE_LARGE
    title.alignment = CENTER
    ws.row_dimensions[1].height = 36

    # ---- k值分组标题行 ----
    k_group_fills = {
        50:  PatternFill("solid", fgColor="1A5276"),
        100: PatternFill("solid", fgColor="145A32"),
        200: PatternFill("solid", fgColor="7D3C98"),
        400: PatternFill("solid", fgColor="784212"),
        500: PatternFill("solid", fgColor="0B5345"),
    }
    model_fills = {'qwen3.5-4B': Qwen_HEAD, 'minicpm-v-4.6': MiniCPM_HEAD}

    # Row 2: k值分组 (merge 2 cols each)
    ws.merge_cells('A2:B2')
    ws['A2'].value = '任务类型'
    style(ws, 2, 1, '任务类型', DARK_BLUE, BOLD_WHITE, CENTER)

    for offset, k in enumerate(K_VALUES):
        col = 3 + offset * 2
        ws.merge_cells(start_row=2, start_column=col,
                       end_row=2, end_column=col + 1)
        c = ws.cell(row=2, column=col, value=f'k={k}')
        c.fill = k_group_fills[k]
        c.font = BOLD_WHITE
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[2].height = 24

    # Row 3: 模型子标题
    ws['A3'].value = '英文名'
    ws['B3'].value = '中文名'
    style(ws, 3, 1, '英文名', DARK_BLUE, BOLD_WHITE, CENTER)
    style(ws, 3, 2, '中文名', DARK_BLUE, BOLD_WHITE, CENTER)
    for offset, k in enumerate(K_VALUES):
        col = 3 + offset * 2
        for mi, model in enumerate(MODELS):
            c = ws.cell(row=3, column=col + mi, value='Qwen' if model == 'qwen3.5-4B' else 'MiniCPM')
            c.fill = model_fills[model]
            c.font = BOLD_WHITE
            c.alignment = CENTER
            c.border = BORDER
    ws.row_dimensions[3].height = 22

    # ---- 数据行 ----
    row = 4
    for ti, qtype in enumerate(TYPE_ORDER):
        row_fill = WHITE_FILL if ti % 2 == 0 else LIGHT_GREY
        ws.cell(row=row, column=1, value=qtype)
        ws.cell(row=row, column=2, value=CHINESE[qtype])
        style(ws, row, 1, qtype, row_fill, NORMAL, LEFT)
        style(ws, row, 2, CHINESE[qtype], row_fill, NORMAL, LEFT)

        for offset, k in enumerate(K_VALUES):
            col_base = 3 + offset * 2
            for mi, model in enumerate(MODELS):
                key = (k, model)
                cell = ws.cell(row=row, column=col_base + mi)
                if key not in all_data:
                    cell.value = 'N/A'
                    style(ws, row, col_base + mi, 'N/A', row_fill, NORMAL, CENTER)
                    continue

                by_type = all_data[key]['by_type']
                s = by_type.get(qtype, {'correct': 0, 'total': 0})
                c_val = s['correct']
                n_val = s['total']
                acc = c_val / n_val if n_val > 0 else None

                if acc is None:
                    cell.value = 'N/A'
                    style(ws, row, col_base + mi, 'N/A', row_fill, NORMAL, CENTER)
                else:
                    cell.value = f'{acc:.1%}'
                    color = ('2E7D32' if acc >= 0.90 else
                             '1565C0' if acc >= 0.70 else
                             'E65100' if acc <  0.50 else '1F1F1F')
                    style(ws, row, col_base + mi, f'{acc:.1%}',
                          row_fill, Font(bold=True, size=10, color=color), CENTER)
        row += 1

    # ---- 合计行 ----
    ws.row_dimensions[row].height = 6  # spacer
    row += 1
    ws.cell(row=row, column=1, value='__total__')
    ws.cell(row=row, column=2, value='合计（L1+L2）')
    style(ws, row, 1, '__total__', DARK_BLUE, BOLD_WHITE, CENTER)
    style(ws, row, 2, '合计（L1+L2）', DARK_BLUE, BOLD_WHITE, LEFT)

    for offset, k in enumerate(K_VALUES):
        col_base = 3 + offset * 2
        for mi, model in enumerate(MODELS):
            key = (k, model)
            if key not in all_data:
                style(ws, row, col_base + mi, 'N/A', DARK_BLUE, BOLD_WHITE, CENTER)
                continue

            overall = all_data[key]['overall']
            c_val = overall['correct']
            n_val = overall['total']
            acc = c_val / n_val if n_val > 0 else None

            if acc is None:
                style(ws, row, col_base + mi, 'N/A', DARK_BLUE, BOLD_WHITE, CENTER)
            else:
                style(ws, row, col_base + mi, f'{acc:.1%}', DARK_BLUE,
                      Font(bold=True, size=11, color='FFFF00'), CENTER)
    row += 1

    # ---- 性能详情行（合并到同一 sheet） ----
    ws.row_dimensions[row].height = 6
    row += 1

    # perf 标题
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    c = ws.cell(row=row, column=1, value='性能指标（每列 = 该 (model, k) 组合的整体统计）')
    c.fill = PERF_HEAD
    c.font = BOLD_WHITE
    c.alignment = CENTER
    row += 1

    # perf 表头
    perf_labels = ['英文名', '中文名',
                   'k=50 Q', 'k=50 M',
                   'k=100 Q', 'k=100 M',
                   'k=200 Q', 'k=200 M',
                   'k=400 Q', 'k=400 M',
                   'k=500 Q', 'k=500 M']
    for ci, lbl in enumerate(perf_labels, start=1):
        c = ws.cell(row=row, column=ci, value=lbl)
        c.fill = PERF_HEAD
        c.font = BOLD_WHITE
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 20
    row += 1

    perf_rows = [
        ('总准确率(L1+L2)', 'total_acc', 3),
        ('平均输出token',   'avg_out',   1),
        ('平均输入token',   'avg_in',    1),
        ('平均图片token',   'avg_img',   1),
        ('截断率(≥4096)',   'trunc_rate',4),
    ]
    for label_cn, key, decimals in perf_rows:
        label_en = key
        row_fill = WHITE_FILL if perf_rows.index((label_cn, key, decimals)) % 2 == 0 else LIGHT_GREY

        ws.cell(row=row, column=1, value=label_en)
        ws.cell(row=row, column=2, value=label_cn)
        style(ws, row, 1, label_en, row_fill, NORMAL, LEFT)
        style(ws, row, 2, label_cn, row_fill, NORMAL, LEFT)

        for offset, k in enumerate(K_VALUES):
            col_base = 3 + offset * 2
            for mi, model in enumerate(MODELS):
                key_data = (k, model)
                cell = ws.cell(row=row, column=col_base + mi)
                if key_data not in all_data:
                    cell.value = 'N/A'
                    style(ws, row, col_base + mi, 'N/A', row_fill, NORMAL, CENTER)
                    continue

                perf = all_data[key_data]['perf']
                if key == 'total_acc':
                    o = all_data[key_data]['overall']
                    v = f'{o["correct"]/o["total"]:.1%}' if o['total'] > 0 else 'N/A'
                elif key == 'avg_out':
                    v = num_str(perf['output_tokens'], decimals)
                elif key == 'avg_in':
                    v = num_str(perf['input_tokens'], decimals)
                elif key == 'avg_img':
                    v = num_str(perf['img_tokens'], decimals)
                elif key == 'trunc_rate':
                    v = pct_str(perf['truncated'], perf['total'])
                style(ws, row, col_base + mi, v, row_fill, NORMAL, CENTER)

        row += 1

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 16
    for i in range(3, 13):
        ws.column_dimensions[get_column_letter(i)].width = 10

    # ---- 冻结前两列 ----
    ws.freeze_panes = 'C4'

    out_path = PROJ / 'Agent_trash' / 'VQA_type_report_v2.xlsx'
    wb.save(str(out_path))
    print(f'已保存: {out_path}')


if __name__ == '__main__':
    main()