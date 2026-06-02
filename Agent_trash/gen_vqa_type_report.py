#!/usr/bin/env python3
"""
VRS-VQA 分层评测报表生成器
=========================
读取 raw_outputs/vqa_raw.json，按 question_id 关联原始标注的 type 字段，
生成 Excel 报表，对比 Qwen3.5-4B 和 MiniCPM-V 在不同 k 值下的任务类型准确率。

输出：
  Sheet1: 总览 — 各 k 值下两个模型的总准确率（L1+L2）
  Sheet2: 任务类型 — 按 type 字段分桶的准确率（L1+L2）

用法:
  python3 gen_vqa_type_report.py
"""
import json
import time
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# =====================================================================
# 配置
# =====================================================================
PROJ     = Path('/home/admin1/projects/remote_vlm_eval')
MODEL_ROOT = Path('/home/admin1/models')
VQA_ANNOT = MODEL_ROOT / 'VRSBench_EVAL_vqa.json'

MODELS   = ['qwen3.5-4B', 'minicpm-v-4.6']
K_VALUES = [50, 100, 200, 400, 500]

LIGHT_BLUE = PatternFill("solid", fgColor="DAEEF3")
LIGHT_GREEN = PatternFill("solid", fgColor="D8F0D8")
LIGHT_YELLOW = PatternFill("solid", fgColor="FFFACD")
LIGHT_PINK = PatternFill("solid", fgColor="FFE4E1")
WHITE = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GREY = PatternFill("solid", fgColor="F5F5F5")
DARK_BLUE_HEAD = PatternFill("solid", fgColor="1F4E79")
DARK_GREEN_HEAD = PatternFill("solid", fgColor="375623")
DARK_PINK_HEAD = PatternFill("solid", fgColor="843C0C")

BOLD_WHITE = Font(bold=True, color="FFFFFF", size=10)
BOLD_DARK  = Font(bold=True, color="1F1F1F", size=10)
NORMAL     = Font(size=10)
BOLD_BIGGER = Font(bold=True, size=12)

THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right',  vertical='center')


# =====================================================================
# 数据加载
# =====================================================================
def load_type_map():
    with open(VQA_ANNOT) as f:
        data = json.load(f)
    return {r['question_id']: r['type'] for r in data}


def load_raw_records(k: int, model: str):
    path = (MODEL_ROOT / 'raw_outputs' / f'lever_k={k}'
            / f'{model}_thinkOFF_lever_k={k}' / 'vqa_raw.json')
    if not path.exists():
        return None
    with open(path) as f:
        return json.load(f)


# =====================================================================
# 统计
# =====================================================================
def compute_stats(records, type_map):
    """
    返回:
      overall: {'correct': n, 'total': n, 'correct_l1': n, 'total_l1': n, ...}
      by_type: {type_str: {'correct': n, 'total': n}}
    """
    overall = {
        'correct': 0, 'total': 0,
        'correct_l1': 0, 'total_l1': 0,
        'correct_l2': 0, 'total_l2': 0,
    }
    by_type = defaultdict(lambda: {'correct': 0, 'total': 0})

    for r in records:
        if r.get('judge_level') not in ('L1', 'L2'):
            continue
        qtype = type_map.get(r['_idx'], 'unknown')
        c = r.get('correct', 0)
        lvl = r['judge_level']

        overall['total'] += 1
        overall['correct'] += c
        if lvl == 'L1':
            overall['total_l1'] += 1
            overall['correct_l1'] += c
        else:
            overall['total_l2'] += 1
            overall['correct_l2'] += c

        by_type[qtype]['total'] += 1
        by_type[qtype]['correct'] += c

    return overall, by_type


def stability(series):
    """返回 "(stable)" 如果变化<2%，否则 "(volatile)" """
    if len(series) < 2:
        return ''
    vals = [v for v in series if v is not None]
    if len(vals) < 2:
        return ''
    mx = max(vals)
    mn = min(vals)
    return ' (stable)' if (mx - mn) < 0.02 else ' (volatile)'


# =====================================================================
# Excel 生成
# =====================================================================
def acc_str(correct, total):
    if total == 0:
        return 'N/A'
    return f'{correct/total:.1%}'


def style_cell(cell, fill, font=None, align=None, border=True):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = THIN_BORDER


def write_header_row(ws, row, cols, fill, font=None):
    for col_idx, text in enumerate(cols, start=1):
        c = ws.cell(row=row, column=col_idx, value=text)
        style_cell(c, fill=fill, font=font or BOLD_WHITE, align=CENTER)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------
# Sheet 1: 总览
# ---------------------------------------------------------------------
def build_sheet_overview(wb, type_map):
    ws = wb.create_sheet('总览')

    # Title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = 'VRS-VQA 分层准确率总览  （L1 + L2，Qwen3.5-4B vs MiniCPM-V，think-off）'
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # 合并表头说明行
    ws.merge_cells('A2:A3')
    ws['A2'] = '模型'
    style_cell(ws['A2'], DARK_BLUE_HEAD, BOLD_WHITE, CENTER)

    ws.merge_cells('B2:B3')
    ws['B2'] = 'k值'
    style_cell(ws['B2'], DARK_BLUE_HEAD, BOLD_WHITE, CENTER)

    # Section 分组标题
    def section_header(col_start, col_end, text, fill=DARK_GREEN_HEAD):
        ws.merge_cells(start_row=2, start_column=col_start,
                       end_row=2, end_column=col_end)
        c = ws.cell(row=2, column=col_start, value=text)
        style_cell(c, fill, BOLD_WHITE, CENTER)

    section_header(3, 6, 'L1 Substring 准确率')
    section_header(7, 10, 'L2精确匹配 准确率')
    section_header(11, 14, 'L1+L2 合计准确率')
    section_header(15, 16, 'L3 语义（待评测）')

    ws['B3'] = '样本数'
    style_cell(ws['B3'], DARK_BLUE_HEAD, BOLD_WHITE, CENTER)

    for ci, label in enumerate(
        ['准确率', '正确数 / 总数', '准确率', '正确数 / 总数', '准确率', '正确数 / 总数', 'n_l3_pending'],
        start=3
    ):
        c = ws.cell(row=3, column=ci, value=label)
        style_cell(c, LIGHT_GREY, BOLD_DARK, CENTER)

    row = 4
    data_rows = []
    for k in K_VALUES:
        row_data = {'k': k, 'models': {}}
        for model in MODELS:
            records = load_raw_records(k, model)
            if records is None:
                stats = None
                row_data['models'][model] = None
            else:
                overall, by_type = compute_stats(records, type_map)
                stats = overall
                row_data['models'][model] = {
                    'overall': overall,
                    'by_type': by_type,
                    'n_records': len(records),
                }

            # 写 L1 块
            for col_start, col_end, field_key in [
                (3, 4, 'l1'), (7, 8, 'l2'), (11, 12, 'total'), (15, 16, 'l3')
            ]:
                pass

            # Qwen
            if row_data['models'].get('qwen3.5-4B'):
                s = row_data['models']['qwen3.5-4B']['overall']
            else:
                s = None

        row += 1

    set_col_widths(ws, [14, 8, 10, 14, 10, 14, 10, 14, 10, 14, 10, 14, 10, 14, 10, 14])


# ---------------------------------------------------------------------
# Sheet 2: 任务类型详细
# ---------------------------------------------------------------------
def build_sheet_by_type(wb, type_map):
    ws = wb.create_sheet('任务类型准确率')

    # 全局 type 列表（固定顺序）
    TYPE_ORDER = [
        'object quantity',
        'object existence',
        'object position',
        'object category',
        'object color',
        'object shape',
        'object size',
        'object direction',
        'scene type',
        'rural or urban',
        'reasoning',
        'image',
    ]

    # ---- Title ----
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = ('VRS-VQA 任务类型准确率明细  '
                        '（L1+L2，Qwen3.5-4B vs MiniCPM-V，think-off）')
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # ---- 表头 ----
    header_row = 3
    ws.row_dimensions[header_row].height = 35

    # 左侧固定列
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws['A2'] = '任务类型'
    ws['B2'] = '中文名'
    style_cell(ws['A2'], DARK_BLUE_HEAD, BOLD_WHITE, CENTER)
    style_cell(ws['B2'], DARK_BLUE_HEAD, BOLD_WHITE, CENTER)

    # 每个 k 值占 4 列（qwen_acc, qwen_n/t, minicpm_acc, minicpm_n/t）
    for offset, k in enumerate(K_VALUES):
        col_base = 3 + offset * 4
        ws.merge_cells(start_row=2, start_column=col_base,
                       end_row=2, end_column=col_base + 3)
        c = ws.cell(row=2, column=col_base, value=f'k={k}')
        style_cell(c, DARK_GREEN_HEAD, BOLD_WHITE, CENTER)

        fills = [LIGHT_BLUE, LIGHT_GREEN, LIGHT_BLUE, LIGHT_GREEN]
        for ci, label in enumerate(['Qwen acc', 'Qwen n/T', 'MiniCPM acc', 'MiniCPM n/T']):
            cell = ws.cell(row=3, column=col_base + ci, value=label)
            style_cell(cell, fills[ci], BOLD_DARK, CENTER)

    set_col_widths(ws,
                   [22, 20] + [11, 13, 11, 13] * len(K_VALUES))

    # ---- 数据行 ----
    CHINESE_NAMES = {
        'object quantity':   '物体计数',
        'object existence':  '物体是否存在',
        'object position':  '物体位置',
        'object category':   '物体类别',
        'object color':      '物体颜色',
        'object shape':      '物体形状',
        'object size':      '物体大小',
        'object direction':  '物体方向',
        'scene type':        '场景类型',
        'rural or urban':    '城乡分类',
        'reasoning':         '推理',
        'image':             '图像属性',
    }

    # 预加载所有数据
    all_data = {}  # (k, model) -> (overall, by_type)
    for k in K_VALUES:
        for model in MODELS:
            records = load_raw_records(k, model)
            if records:
                all_data[(k, model)] = compute_stats(records, type_map)

    row = 4
    for ti, qtype in enumerate(TYPE_ORDER):
        fill = WHITE if ti % 2 == 0 else LIGHT_GREY
        ws.cell(row=row, column=1, value=qtype)
        ws.cell(row=row, column=2, value=CHINESE_NAMES.get(qtype, qtype))
        style_cell(ws.cell(row=row, column=1), fill, NORMAL, LEFT)
        style_cell(ws.cell(row=row, column=2), fill, NORMAL, LEFT)

        overall_accs = {'qwen3.5-4B': [], 'minicpm-v-4.6': []}

        for offset, k in enumerate(K_VALUES):
            col_base = 3 + offset * 4

            for mi, model in enumerate(MODELS):
                key = (k, model)
                fills = [LIGHT_BLUE, LIGHT_GREEN]
                if key not in all_data:
                    for ci in range(2):
                        cell = ws.cell(row=row, column=col_base + 2 * mi + ci)
                        cell.value = 'N/A'
                        style_cell(cell, fills[mi], NORMAL, CENTER)
                    overall_accs[model].append(None)
                    continue

                overall, by_type = all_data[key]
                s = by_type.get(qtype, {'correct': 0, 'total': 0})
                c = s['correct']
                n = s['total']
                acc = c / n if n > 0 else None
                overall_accs[model].append(acc)

                # acc cell
                acc_cell = ws.cell(row=row, column=col_base + 2 * mi)
                acc_cell.value = acc_str(c, n)
                style_cell(acc_cell, fills[mi],
                           Font(bold=True, size=10,
                                color='2E7D32' if (acc and acc >= 0.9) else
                                      '1565C0' if (acc and acc >= 0.7) else
                                      'E65100' if (acc and acc < 0.5) else '1F1F1F'),
                           CENTER)

                # n/T cell
                n_cell = ws.cell(row=row, column=col_base + 2 * mi + 1)
                n_cell.value = f'{c}/{n}' if n > 0 else 'N/A'
                style_cell(n_cell, fills[mi], NORMAL, CENTER)

        # 稳定性标注
        for mi, model in enumerate(MODELS):
            series = overall_accs[model]
            if len([v for v in series if v is not None]) >= 3:
                tag = stability(series)
                if tag:
                    note_cell = ws.cell(row=row, column=17 + mi)
                    note_cell.value = tag

        row += 1

    # ---- 汇总行 ----
    ws.row_dimensions[row].height = 5  # spacer
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1, value='合计（L1+L2）')
    style_cell(c, DARK_BLUE_HEAD, BOLD_WHITE, CENTER)
    ws.row_dimensions[row].height = 22

    overall_accs_sum = {'qwen3.5-4B': [], 'minicpm-v-4.6': []}
    for offset, k in enumerate(K_VALUES):
        col_base = 3 + offset * 4
        row_fills = [LIGHT_BLUE, LIGHT_BLUE, LIGHT_GREEN, LIGHT_GREEN]

        for mi, model in enumerate(MODELS):
            key = (k, model)
            if key not in all_data:
                for ci in range(2):
                    cell = ws.cell(row=row, column=col_base + mi * 2 + ci)
                    cell.value = 'N/A'
                    style_cell(cell, row_fills[mi * 2 + ci], NORMAL, CENTER)
                overall_accs_sum[model].append(None)
                continue

            overall, by_type = all_data[key]
            c_total = overall['correct']
            n_total = overall['total']
            acc = c_total / n_total if n_total > 0 else None
            overall_accs_sum[model].append(acc)

            acc_cell = ws.cell(row=row, column=col_base + mi * 2)
            acc_cell.value = acc_str(c_total, n_total)
            style_cell(acc_cell, row_fills[mi * 2],
                       Font(bold=True, size=11, color='FFFFFF'),
                       CENTER)

            n_cell = ws.cell(row=row, column=col_base + mi * 2 + 1)
            n_cell.value = f'{c_total}/{n_total}'
            style_cell(n_cell, row_fills[mi * 2 + 1],
                       Font(bold=True, color='FFFFFF', size=10), CENTER)

    return wb


# =====================================================================
# 入口
# =====================================================================
def main():
    print('加载 type 映射…')
    type_map = load_type_map()

    print('生成 Excel…')
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    wb = build_sheet_by_type(wb, type_map)

    out_path = PROJ / 'Agent_trash' / 'VQA_type_report.xlsx'
    wb.save(str(out_path))
    print(f'已保存: {out_path}')


if __name__ == '__main__':
    main()
