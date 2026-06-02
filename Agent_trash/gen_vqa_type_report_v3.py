#!/usr/bin/env python3
"""
VRS-VQA 分层评测报表生成器 v3（行列互换）
========================================
每行 = 一个 (模型, k值) 组合
列  =  指标名（模型|k|总准确率|平均输出token|平均输入token|平均图片token|截断率|12种任务类型准确率）

用法:
  python3 gen_vqa_type_report_v3.py
"""
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================================
# 配置
# =====================================================================
PROJ       = Path('/home/admin1/projects/remote_vlm_eval')
MODEL_ROOT = Path('/home/admin1/models')
VQA_ANNOT  = MODEL_ROOT / 'VRSBench_EVAL_vqa.json'

MODELS   = ['qwen3.5-4B', 'minicpm-v-4.6']
K_VALUES = [50, 100, 200, 400, 500]

# 列定义（固定顺序）
COL_LABELS = [
    '模型', 'k值',
    '总准确率\n(L1+L2)', '平均输出\ntoken', '平均输入\ntoken',
    '平均图片\ntoken', '截断率\n(≥4096)',
    'object\nquantity', 'object\nexistence', 'object\nposition',
    'object\ncategory', 'object\ncolor', 'object\nshape',
    'object\nsize', 'object\ndirection', 'scene\ntype',
    'rural or\nurban', 'reasoning', 'image',
]
N_COLS = len(COL_LABELS)  # 20

# 任务类型（对应 col 8-19）
TYPE_ORDER = [
    'object quantity', 'object existence', 'object position',
    'object category', 'object color', 'object shape',
    'object size', 'object direction', 'scene type',
    'rural or urban', 'reasoning', 'image',
]
CHINESE = {
    'object quantity':  '物体计数',
    'object existence': '物体是否存在',
    'object position':  '物体位置',
    'object category':   '物体类别',
    'object color':      '物体颜色',
    'object shape':      '物体形状',
    'object size':      '物体大小',
    'object direction':  '物体方向',
    'scene type':        '场景类型',
    'rural or urban':    '城乡分类',
    'reasoning':         '推理',
    'image':             '图像属性',
}

# =====================================================================
# 颜色
# =====================================================================
# 行分组底色（每个 k 一组，两行 Qwen+MiniCPM）
K_FILLS = {
    50:  PatternFill("solid", fgColor="D6EAF8"),
    100: PatternFill("solid", fgColor="D5F5E3"),
    200: PatternFill("solid", fgColor="FAD7A0"),
    400: PatternFill("solid", fgColor="D7BDE2"),
    500: PatternFill("solid", fgColor="ABEBC6"),
}
MODEL_FILLS = {
    'qwen3.5-4B':   PatternFill("solid", fgColor="2E86AB"),
    'minicpm-v-4.6': PatternFill("solid", fgColor="1D8348"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
TOTAL_FILL  = PatternFill("solid", fgColor="17202A")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

BOLD_WHITE = Font(bold=True, color="FFFFFF", size=10)
BOLD_DARK  = Font(bold=True, color="1F1F1F", size=10)
NORMAL     = Font(size=10)
BOLD_TITLE = Font(bold=True, color="FFFFFF", size=14)

THIN   = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=False)

# =====================================================================
# 数据加载
# =====================================================================
def load_type_map():
    with open(VQA_ANNOT) as f:
        return {r['question_id']: r['type'] for r in json.load(f)}


def load_records(k, model):
    path = MODEL_ROOT / 'raw_outputs' / f'lever_k={k}' / f'{model}_thinkOFF_lever_k={k}' / 'vqa_raw.json'
    return json.load(open(path)) if path.exists() else None


def compute_stats(records, type_map):
    by_type  = defaultdict(lambda: {'correct': 0, 'total': 0})
    overall  = {'correct': 0, 'total': 0}
    perf     = {'out': [], 'in': [], 'img': [], 'trunc': 0, 'n': 0}

    for r in records:
        lvl = r.get('judge_level', '')
        if lvl not in ('L1', 'L2'):
            continue
        qtype = type_map.get(r['_idx'], 'unknown')
        c     = r.get('correct', 0)

        overall['total']    += 1
        overall['correct']  += c
        by_type[qtype]['total']    += 1
        by_type[qtype]['correct']  += c

        perf['out'].append(r.get('tokens', 0))
        perf['in'].append(r['input_tokens'])
        perf['img'].append(r['img_tokens'])
        if r.get('tokens', 0) >= 4096:
            perf['trunc'] += 1
        perf['n'] += 1

    return by_type, overall, perf


# =====================================================================
# 工具
# =====================================================================
def fmt_pct(c, n):
    return f'{c/n:.1%}' if n > 0 else 'N/A'

def fmt_num(vals, dec=1):
    return f'{sum(vals)/len(vals):.{dec}f}' if vals else 'N/A'

def style(ws, row, col, value, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:  c.fill = fill
    if font:  c.font = font
    if align: c.alignment = align
    c.border = BORDER
    return c

def acc_color(acc):
    if acc is None: return '1F1F1F'
    return '2E7D32' if acc >= 0.90 else '1565C0' if acc >= 0.70 else 'E65100' if acc < 0.50 else '1F1F1F'


# =====================================================================
# 主函数
# =====================================================================
def main():
    print('加载 type 映射…')
    type_map = load_type_map()

    print('加载评测记录…')
    all_data = {}
    for k in K_VALUES:
        for model in MODELS:
            recs = load_records(k, model)
            if recs is None:
                continue
            by_type, overall, perf = compute_stats(recs, type_map)
            all_data[(k, model)] = {'by_type': by_type, 'overall': overall, 'perf': perf}
    print(f'已加载 {len(all_data)} 个组合')

    # ---- 构建 Excel ----
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('VRS-VQA任务类型准确率')

    # ---- 标题 ----
    ws.merge_cells(f'A1:{get_column_letter(N_COLS)}1')
    tc = ws['A1']
    tc.value = 'VRS-VQA 任务类型准确率明细  (Qwen3.5-4B vs MiniCPM-V 4.6, think-off)'
    tc.fill  = HEADER_FILL
    tc.font  = BOLD_TITLE
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 38

    # ---- 列标题行 ----
    ws.row_dimensions[2].height = 50
    for ci, label in enumerate(COL_LABELS, start=1):
        c = ws.cell(row=2, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = BOLD_WHITE
        c.alignment = CENTER
        c.border = BORDER

    # ---- 数据行：先 Qwen 全量，再 MiniCPM 全量 ----
    row = 3
    for model in MODELS:
        for k in K_VALUES:
            key = (k, model)
            kfill = K_FILLS.get(k, WHITE_FILL)
            mfill = MODEL_FILLS.get(model, kfill)

            # 模型名（col 1）
            style(ws, row, 1, 'Qwen3.5-4B' if model == 'qwen3.5-4B' else 'MiniCPM-V 4.6',
                  kfill, Font(bold=True, size=10, color='FFFFFF'), LEFT)
            # k值（col 2）
            style(ws, row, 2, k, kfill, Font(bold=True, size=10, color='FFFFFF'), CENTER)

            if key not in all_data:
                for ci in range(3, N_COLS + 1):
                    style(ws, row, ci, 'N/A', kfill, NORMAL, CENTER)
                row += 1
                continue

            d   = all_data[key]
            o   = d['overall']
            p   = d['perf']
            bt  = d['by_type']

            # col 3: 总准确率
            acc_total = o['correct'] / o['total'] if o['total'] > 0 else None
            c3 = ws.cell(row=row, column=3, value=fmt_pct(o['correct'], o['total']))
            c3.fill = kfill
            c3.font = Font(bold=True, size=11, color=acc_color(acc_total))
            c3.alignment = CENTER
            c3.border = BORDER

            # col 4: 平均输出token
            style(ws, row, 4, fmt_num(p['out'], 1), kfill, NORMAL, CENTER)

            # col 5: 平均输入token
            style(ws, row, 5, fmt_num(p['in'], 0), kfill, NORMAL, CENTER)

            # col 6: 平均图片token
            style(ws, row, 6, fmt_num(p['img'], 0), kfill, NORMAL, CENTER)

            # col 7: 截断率
            trunc_rate = p['trunc'] / p['n'] if p['n'] > 0 else 0
            c7 = ws.cell(row=row, column=7, value=f'{trunc_rate:.2%}')
            c7.fill = kfill
            c7.font = Font(size=10, color='E65100' if trunc_rate > 0.1 else '1F1F1F')
            c7.alignment = CENTER
            c7.border = BORDER

            # col 8-19: 12种任务类型准确率
            for ti, qtype in enumerate(TYPE_ORDER):
                col = 8 + ti
                s   = bt.get(qtype, {'correct': 0, 'total': 0})
                c_v = s['correct']
                n_v = s['total']
                acc = c_v / n_v if n_v > 0 else None

                cell = ws.cell(row=row, column=col,
                               value=fmt_pct(c_v, n_v) if acc is not None else 'N/A')
                cell.fill   = kfill
                cell.font   = Font(bold=True, size=10, color=acc_color(acc))
                cell.alignment = CENTER
                cell.border = BORDER

            row += 1

    # ---- 列宽 ----
    ws.column_dimensions['A'].width = 14   # 模型
    ws.column_dimensions['B'].width = 8    # k值
    for ci in range(3, N_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.column_dimensions[get_column_letter(N_COLS)].width = 10  # image

    # ---- 冻结前两列 ----
    ws.freeze_panes = 'C3'

    # ---- 保存 ----
    out_path = PROJ / 'Agent_trash' / 'VQA_type_report_v3.xlsx'
    wb.save(str(out_path))
    print(f'已保存: {out_path}')


if __name__ == '__main__':
    main()