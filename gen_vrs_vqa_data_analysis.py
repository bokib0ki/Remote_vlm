"""
VRS-VQA 数据集统计分析脚本
读取 /home/admin1/models/VRSBench_EVAL_vqa.json，统计:
  - 总体平均答案长度
  - 单词/短语/句子比例（全局 + 各子类型）
  - 各子类型最高频答案
输出: /home/admin1/projects/remote_vlm_eval/excel_result/VRS-VQA-DataAnalysis.excel

分类规则:
  sentence: 答案含 . ? ! 句末标点
  word:     单词/数字（无空格无标点）
  phrase:   多个词组成，但非完整句子

长度:
  avg_word_len: 答案按空白分词后的词数
  avg_char_len: 答案字符数
"""

import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------- 路径配置 ----------
SRC_JSON = Path('/home/admin1/models/VRSBench_EVAL_vqa.json')
OUT_DIR = Path('/home/admin1/projects/remote_vlm_eval/excel_result')
OUT_PATH = OUT_DIR / 'VRS-VQA-DataAnalysis.excel'

# 12 type 固定顺序（按数据量从大到小，type 名沿用 VRSBench 原文）
TYPE_ORDER = [
    'object existence',
    'object quantity',
    'object position',
    'object category',
    'object color',
    'scene type',
    'object shape',
    'image',
    'object size',
    'reasoning',
    'object direction',
    'rural or urban',
]

# ---------- 分类函数 ----------
def classify_answer(ans: str) -> str:
    s = ans.strip()
    if not s:
        return 'word'  # 空字符串视作单词，统计上几乎不出现
    if re.search(r'[.?!]', s):
        return 'sentence'
    if ' ' not in s and '\t' not in s:
        return 'word'
    return 'phrase'

def word_count(ans: str) -> int:
    return len(ans.split())

# ---------- 统计 ----------
def analyze(records):
    total = len(records)

    # 全局
    overall = {
        'total': total,
        'cat_counter': Counter(),
        'char_sum': 0,
        'word_sum': 0,
    }
    # per type
    per_type = {}
    for t in TYPE_ORDER:
        per_type[t] = {
            'n': 0,
            'cat_counter': Counter(),
            'char_sum': 0,
            'word_sum': 0,
            'gt_counter': Counter(),
        }

    for d in records:
        t = d['type']
        gt = d['ground_truth']
        cat = classify_answer(gt)
        wc = word_count(gt)
        char = len(gt)

        overall['cat_counter'][cat] += 1
        overall['char_sum'] += char
        overall['word_sum'] += wc

        if t not in per_type:
            # 防御性: 数据中如果有未列出的 type, 动态加
            per_type[t] = {
                'n': 0, 'cat_counter': Counter(),
                'char_sum': 0, 'word_sum': 0, 'gt_counter': Counter(),
            }
        s = per_type[t]
        s['n'] += 1
        s['cat_counter'][cat] += 1
        s['char_sum'] += char
        s['word_sum'] += wc
        s['gt_counter'][gt] += 1

    return overall, per_type


# ---------- Excel 样式 ----------
HEADER_FILL = PatternFill('solid', fgColor='2F5597')   # 深蓝
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBHEAD_FILL = PatternFill('solid', fgColor='D9E1F2')  # 浅蓝
SUBHEAD_FONT = Font(bold=True, color='1F3864', size=10)
TOTAL_FILL = PatternFill('solid', fgColor='FFE699')    # 浅黄
TOTAL_FONT = Font(bold=True, color='806000', size=10)
TOP1_FILL = PatternFill('solid', fgColor='FFD966')     # 最高频答案底色
THIN = Side(border_style='thin', color='B4B4B4')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, row_start, row_end, ncols, num_cols=()):
    """应用边框 + 数值居中 + 文本左对齐"""
    for r in range(row_start, row_end + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if c in num_cols:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Sheet 1: 总体统计 ----------
def write_overall_sheet(wb, overall):
    ws = wb.create_sheet('总体统计')
    total = overall['total']
    n_word = overall['cat_counter'].get('word', 0)
    n_phrase = overall['cat_counter'].get('phrase', 0)
    n_sent = overall['cat_counter'].get('sentence', 0)
    avg_word = overall['word_sum'] / total
    avg_char = overall['char_sum'] / total

    rows = [
        ('总样本数', total, ''),
        ('平均答案长度（词数）', round(avg_word, 3), 'word_count(answer)'),
        ('平均答案长度（字符数）', round(avg_char, 3), 'len(answer)'),
        ('', '', ''),
        ('分类', '数量', '占比'),
        ('单词 (word)', n_word, f'{n_word/total*100:.2f}%'),
        ('短语 (phrase)', n_phrase, f'{n_phrase/total*100:.2f}%'),
        ('句子 (sentence)', n_sent, f'{n_sent/total*100:.2f}%'),
    ]

    # 写标题
    ws.cell(row=1, column=1, value='VRS-VQA 数据集 — 总体统计')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    # 子表头（行 3）
    for c, h in enumerate(['指标', '值', '说明'], start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 3)

    for i, (k, v, note) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=3, value=note)
    # 分类段着色
    for r in range(4, 8):
        ws.cell(row=r, column=1).fill = SUBHEAD_FILL
        ws.cell(row=r, column=1).font = SUBHEAD_FONT
    style_body(ws, 4, 11, 3, num_cols=(2,))

    set_col_widths(ws, [28, 18, 32])
    ws.row_dimensions[1].height = 24

    # 顶部统计 (行 4-6) 高亮
    for r in (4, 5, 6):
        ws.cell(row=r, column=1).fill = TOTAL_FILL
        ws.cell(row=r, column=1).font = TOTAL_FONT


# ---------- Sheet 2: 各子类型统计 ----------
def write_per_type_sheet(wb, per_type):
    ws = wb.create_sheet('各子类型统计')

    headers = [
        'Type',
        'N (题数)',
        '最高频答案',
        '最高频答案出现次数',
        '最高频答案占比',
        '平均答案长度(词数)',
        '平均答案长度(字符数)',
        '单词数',
        '单词比例',
        '短语数',
        '短语比例',
        '句子数',
        '句子比例',
    ]
    # 标题
    ws.cell(row=1, column=1, value='VRS-VQA 数据集 — 各子类型统计')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # 表头
    header_row = 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))

    # 数据行
    row = header_row + 1
    type_rows = []
    for t in TYPE_ORDER:
        if t not in per_type:
            continue
        s = per_type[t]
        n = s['n']
        gt_top, gt_top_n = s['gt_counter'].most_common(1)[0]
        gt_top_pct = gt_top_n / n * 100
        avg_w = s['word_sum'] / n
        avg_c = s['char_sum'] / n
        n_w = s['cat_counter'].get('word', 0)
        n_p = s['cat_counter'].get('phrase', 0)
        n_s = s['cat_counter'].get('sentence', 0)
        type_rows.append((t, n, gt_top, gt_top_n, gt_top_pct, avg_w, avg_c, n_w, n_p, n_s))

        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=2, value=n)
        ws.cell(row=row, column=3, value=gt_top)
        ws.cell(row=row, column=4, value=gt_top_n)
        ws.cell(row=row, column=5, value=f'{gt_top_pct:.2f}%')
        ws.cell(row=row, column=6, value=round(avg_w, 3))
        ws.cell(row=row, column=7, value=round(avg_c, 3))
        ws.cell(row=row, column=8, value=n_w)
        ws.cell(row=row, column=9, value=f'{n_w/n*100:.2f}%')
        ws.cell(row=row, column=10, value=n_p)
        ws.cell(row=row, column=11, value=f'{n_p/n*100:.2f}%')
        ws.cell(row=row, column=12, value=n_s)
        ws.cell(row=row, column=13, value=f'{n_s/n*100:.2f}%')

        # 最高频答案 cell 高亮
        ws.cell(row=row, column=3).fill = TOP1_FILL
        ws.cell(row=row, column=3).font = Font(bold=True, color='806000')
        row += 1

    # 合计行
    total_n = sum(s['n'] for s in per_type.values())
    total_w = sum(s['cat_counter'].get('word', 0) for s in per_type.values())
    total_p = sum(s['cat_counter'].get('phrase', 0) for s in per_type.values())
    total_s = sum(s['cat_counter'].get('sentence', 0) for s in per_type.values())
    total_word_sum = sum(s['word_sum'] for s in per_type.values())
    total_char_sum = sum(s['char_sum'] for s in per_type.values())

    ws.cell(row=row, column=1, value='TOTAL (全部 type)')
    ws.cell(row=row, column=2, value=total_n)
    ws.cell(row=row, column=3, value='—')
    ws.cell(row=row, column=4, value='—')
    ws.cell(row=row, column=5, value='—')
    ws.cell(row=row, column=6, value=round(total_word_sum / total_n, 3))
    ws.cell(row=row, column=7, value=round(total_char_sum / total_n, 3))
    ws.cell(row=row, column=8, value=total_w)
    ws.cell(row=row, column=9, value=f'{total_w/total_n*100:.2f}%')
    ws.cell(row=row, column=10, value=total_p)
    ws.cell(row=row, column=11, value=f'{total_p/total_n*100:.2f}%')
    ws.cell(row=row, column=12, value=total_s)
    ws.cell(row=row, column=13, value=f'{total_s/total_n*100:.2f}%')

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT

    last_data_row = row
    style_body(ws, header_row + 1, last_data_row, len(headers),
               num_cols=(2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13))

    set_col_widths(ws, [22, 9, 26, 14, 14, 16, 18, 10, 12, 10, 12, 10, 12])
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A4'


# ---------- Sheet 3: 各 type 答案 Top-10 频次 ----------
def write_top_answers_sheet(wb, per_type):
    ws = wb.create_sheet('各子类型Top10答案')

    ws.cell(row=1, column=1, value='VRS-VQA — 各子类型 Top-10 高频答案')
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    headers = ['Type', '排名', '答案', '出现次数']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 4)

    row = 4
    for t in TYPE_ORDER:
        if t not in per_type:
            continue
        s = per_type[t]
        # type 标题行（合并 A 列）
        start_row = row
        ws.cell(row=row, column=1, value=t)
        ws.cell(row=row, column=1).fill = SUBHEAD_FILL
        ws.cell(row=row, column=1).font = SUBHEAD_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = BORDER
        # 把标题行的 2-4 列填色（避免合并后样式缺失）
        for c in (2, 3, 4):
            ws.cell(row=row, column=c).fill = SUBHEAD_FILL
            ws.cell(row=row, column=c).border = BORDER
        row += 1
        # top10
        for rank, (ans, cnt) in enumerate(s['gt_counter'].most_common(10), start=1):
            ws.cell(row=row, column=1, value='')
            ws.cell(row=row, column=2, value=rank)
            ws.cell(row=row, column=3, value=ans)
            ws.cell(row=row, column=4, value=cnt)
            if rank == 1:
                ws.cell(row=row, column=3).fill = TOP1_FILL
                ws.cell(row=row, column=3).font = Font(bold=True, color='806000')
            row += 1
        # 合并 type 标题的 A 列
        if row - 1 > start_row:
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=row - 1, end_column=1)
            ws.cell(row=start_row, column=1).alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    last_data_row = row - 1
    style_body(ws, 4, last_data_row, 4, num_cols=(2, 4))

    set_col_widths(ws, [22, 8, 38, 14])
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A4'


# ---------- main ----------
def main():
    if not SRC_JSON.exists():
        raise SystemExit(f'源数据不存在: {SRC_JSON}')

    with open(SRC_JSON, encoding='utf-8') as f:
        records = json.load(f)
    print(f'加载 {len(records)} 条记录 from {SRC_JSON}')

    overall, per_type = analyze(records)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    write_overall_sheet(wb, overall)
    write_per_type_sheet(wb, per_type)
    write_top_answers_sheet(wb, per_type)

    wb.save(OUT_PATH)

    # 控制台汇报
    print(f'\n=== 全局 ({overall["total"]} 条) ===')
    print(f'  avg 词数 : {overall["word_sum"]/overall["total"]:.3f}')
    print(f'  avg 字符 : {overall["char_sum"]/overall["total"]:.3f}')
    print(f'  分类     : {dict(overall["cat_counter"])}')

    print('\n=== 各子类型 ===')
    for t in TYPE_ORDER:
        if t in per_type:
            s = per_type[t]
            n = s['n']
            top, top_n = s['gt_counter'].most_common(1)[0]
            print(f'  {t:25s} N={n:5d}  top="{top}" ({top_n}/{n}={top_n/n*100:.1f}%)  '
                  f'w={s["cat_counter"]["word"]} p={s["cat_counter"]["phrase"]} '
                  f's={s["cat_counter"]["sentence"]}  avg_w={s["word_sum"]/n:.2f} '
                  f'avg_c={s["char_sum"]/n:.2f}')

    print(f'\n✅ Excel 已保存: {OUT_PATH}')


if __name__ == '__main__':
    main()
